import os
import io
import re
import sqlite3
import time
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import APIRouter
import openpyxl

import crypto

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data.db")

BACKUP_PROFIT = os.path.join(BASE_DIR, "backup", "douyin-profit-tracker", "data.db")
BACKUP_DASHBOARD = os.path.join(BASE_DIR, "backup", "shop-dashboard", "data.db")
BACKUP_TASK = os.path.join(BASE_DIR, "backup", "task-tracker", "data.db")

REPORT_ROOT = os.path.dirname(BASE_DIR)

app = FastAPI()
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

profit_router = APIRouter(prefix="/profit")
dashboard_router = APIRouter(prefix="/dashboard")
task_router = APIRouter(prefix="/task")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def calc_profit(douyin_amount, taobao_amount, refund_status, is_influencer, warehouse_status):
    profit = douyin_amount - taobao_amount
    if refund_status in ('已发货退款', '退货退款'):
        profit -= 6.9
    elif refund_status not in ('待发货',) and warehouse_status == '未到仓库':
        pass
    elif refund_status not in ('待发货',) and warehouse_status == '已到达仓库未发货':
        pass
    else:
        profit -= 6.9
    if is_influencer == '是' and refund_status in ('待发货', '已发货'):
        profit -= douyin_amount * 0.25
    if warehouse_status in ('已到达仓库未发货', '已到仓库已发货'):
        if refund_status in ('待发货', '已发货'):
            profit -= 5.5
    if warehouse_status == '已到达仓库未发货' and refund_status not in ('待发货', '已发货'):
        profit -= 3.5
    elif refund_status in ('已发货退款', '退货退款'):
        profit -= 3.5
    return profit


def safe_float(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    try:
        if isinstance(v, str):
            v = v.replace('元', '').replace('/单起', '').replace('/单', '').strip()
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_int(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def extract_date_from_filename(filename):
    m = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', filename)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def scan_excel_files():
    files = []
    if not os.path.exists(REPORT_ROOT):
        return files
    for entry in os.listdir(REPORT_ROOT):
        subdir = os.path.join(REPORT_ROOT, entry)
        if os.path.isdir(subdir) and re.match(r'\d{4}-\d{2}', entry):
            for fname in os.listdir(subdir):
                if fname.endswith('.xlsx') and not fname.startswith('~'):
                    fpath = os.path.join(subdir, fname)
                    d = extract_date_from_filename(fname)
                    if d:
                        files.append((fpath, d))
    return files


def import_excel(filepath, snapshot_date):
    conn = get_db()
    existing = conn.execute("SELECT COUNT(*) FROM daily_snapshot WHERE snapshot_date=?", (snapshot_date,)).fetchone()[0]
    if existing > 0:
        conn.close()
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value or '').strip())

    col_map = {}
    for i, h in enumerate(headers):
        if h == '店铺名称': col_map['shop_name'] = i
        elif h == '状态': col_map['status'] = i
        elif h == '到期时间': col_map['expire_info'] = i
        elif h == '预警': col_map['warning'] = i
        elif h == '成交金额': col_map['amount'] = i
        elif h == '订单数': col_map['order_count'] = i
        elif h == '曝光人数': col_map['exposure'] = i
        elif h == '点击人数': col_map['clicks'] = i
        elif h == '退款金额': col_map['refund_amount'] = i
        elif h == '体验分': col_map['experience_score'] = i
        elif h == '待发货': col_map['pending_ship'] = i
        elif h == '待售后': col_map['pending_after_sale'] = i
        elif h == '在售中': col_map['on_sale'] = i
        elif h == '仓库中': col_map['in_warehouse'] = i
        elif h == '草稿箱': col_map['draft'] = i
        elif h == '24h需发货': col_map['ship_24h'] = i
        elif h == '超时未发货': col_map['overdue_ship'] = i
        elif h == '24h需售后': col_map['after_sale_24h'] = i
        elif h == '待整改': col_map['pending_rectify'] = i
        elif h == '最近违规时间': col_map['last_violation'] = i
        elif h == '工单': col_map['work_order'] = i
        elif h == '资金': col_map['fund'] = i
        elif h == '体验金': col_map['experience_fund'] = i
        elif h == '运费险价格': col_map['insurance_price'] = i
        elif h == '30天有效订单': col_map['valid_orders_30d'] = i
        elif h == '被清洗订单': col_map['cleaned_orders'] = i
        elif h == '违规数': col_map['violation_count'] = i
        elif h == '商品分': col_map['product_score'] = i
        elif h == '物流分': col_map['logistics_score'] = i
        elif h == '服务分': col_map['service_score'] = i
        elif h == '异常包裹': col_map['abnormal_parcel'] = i
        elif h == '回收站': col_map['recycle_bin'] = i
        elif h == '退款金额（今日支付）': col_map['refund_today'] = i
        elif h == '店铺ID': col_map['shop_id_col'] = i
        elif h == '千川ID': col_map['qianchuan_id'] = i
        elif h == '需开票': col_map['need_invoice'] = i
        elif h == '今订单': col_map['today_orders'] = i
        elif h == '分组': col_map['group_name'] = i
        elif h == '备注': col_map['remark'] = i

    def g(row, key, default=''):
        idx = col_map.get(key)
        if idx is not None and idx < len(row) and row[idx] is not None:
            return row[idx]
        return default

    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        shop_name = str(g(row, 'shop_name', '')).strip()
        if not shop_name:
            continue

        sid_col = g(row, 'shop_id_col', '')
        shop_id_str = str(sid_col).strip() if sid_col else ''
        group = str(g(row, 'group_name', '')).strip()

        existing_shop = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
        if existing_shop:
            db_shop_id = existing_shop['id']
            conn.execute("UPDATE shops SET shop_id=?, group_name=? WHERE id=?", (shop_id_str, group, db_shop_id))
        else:
            cur = conn.execute("INSERT INTO shops(name, shop_id, group_name) VALUES(?,?,?)", (shop_name, shop_id_str, group))
            db_shop_id = cur.lastrowid

        conn.execute("""INSERT OR REPLACE INTO daily_snapshot(
            shop_id, snapshot_date, status, expire_info, warning, amount, order_count,
            exposure, clicks, refund_amount, experience_score, pending_ship,
            pending_after_sale, on_sale, in_warehouse, draft, ship_24h, overdue_ship,
            after_sale_24h, pending_rectify, last_violation, work_order, fund,
            experience_fund, insurance_price, valid_orders_30d, cleaned_orders,
            violation_count, product_score, logistics_score, service_score,
            abnormal_parcel, recycle_bin, refund_today, qianchuan_id, need_invoice, today_orders, remark
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (db_shop_id, snapshot_date,
             str(g(row, 'status', '')),
             str(g(row, 'expire_info', '')),
             safe_int(g(row, 'warning', 0)),
             safe_float(g(row, 'amount', 0)),
             safe_int(g(row, 'order_count', 0)),
             safe_int(g(row, 'exposure', 0)),
             safe_int(g(row, 'clicks', 0)),
             safe_float(g(row, 'refund_amount', 0)),
             str(g(row, 'experience_score', '')),
             safe_int(g(row, 'pending_ship', 0)),
             safe_int(g(row, 'pending_after_sale', 0)),
             safe_int(g(row, 'on_sale', 0)),
             safe_int(g(row, 'in_warehouse', 0)),
             safe_int(g(row, 'draft', 0)),
             str(g(row, 'ship_24h', '')),
             safe_int(g(row, 'overdue_ship', 0)),
             str(g(row, 'after_sale_24h', '')),
             safe_int(g(row, 'pending_rectify', 0)),
             str(g(row, 'last_violation', '')),
             safe_int(g(row, 'work_order', 0)),
             safe_float(g(row, 'fund', 0)),
             safe_float(g(row, 'experience_fund', 0)),
             str(g(row, 'insurance_price', '')),
             safe_int(g(row, 'valid_orders_30d', 0)),
             safe_int(g(row, 'cleaned_orders', 0)),
             safe_int(g(row, 'violation_count', 0)),
             str(g(row, 'product_score', '')),
             str(g(row, 'logistics_score', '')),
             str(g(row, 'service_score', '')),
             safe_int(g(row, 'abnormal_parcel', 0)),
             safe_int(g(row, 'recycle_bin', 0)),
             safe_float(g(row, 'refund_today', 0)),
             str(g(row, 'qianchuan_id', '')),
             safe_float(g(row, 'need_invoice', 0)),
             safe_int(g(row, 'today_orders', 0)),
             str(g(row, 'remark', ''))))
        cnt += 1

    conn.commit()
    conn.close()
    wb.close()
    return cnt


def auto_import_all():
    for fpath, d in scan_excel_files():
        import_excel(fpath, d)


def get_summary(conn, snapshot_date):
    return dict(conn.execute("""
        SELECT COUNT(*) as shop_count,
               COALESCE(SUM(amount),0) as total_amount,
               COALESCE(SUM(order_count),0) as total_orders,
               COALESCE(SUM(refund_amount),0) as total_refund,
               COALESCE(SUM(refund_today),0) as total_refund_today,
               COALESCE(SUM(pending_ship),0) as total_pending_ship,
               COALESCE(SUM(overdue_ship),0) as total_overdue_ship,
               COALESCE(SUM(exposure),0) as total_exposure,
               COALESCE(SUM(clicks),0) as total_clicks,
               COALESCE(SUM(order_count),0) as total_today_orders
        FROM daily_snapshot WHERE snapshot_date=?
    """, (snapshot_date,)).fetchone())


def get_range_summary(conn, start_date, end_date):
    return dict(conn.execute("""
        SELECT COUNT(DISTINCT shop_id) as shop_count,
               COALESCE(SUM(amount),0) as total_amount,
               COALESCE(SUM(order_count),0) as total_orders,
               COALESCE(SUM(refund_amount),0) as total_refund,
               COALESCE(SUM(refund_today),0) as total_refund_today,
               COALESCE(SUM(pending_ship),0) as total_pending_ship,
               COALESCE(SUM(overdue_ship),0) as total_overdue_ship,
               COALESCE(SUM(exposure),0) as total_exposure,
               COALESCE(SUM(clicks),0) as total_clicks,
               COALESCE(SUM(order_count),0) as total_today_orders
        FROM daily_snapshot WHERE snapshot_date >= ? AND snapshot_date <= ?
    """, (start_date, end_date)).fetchone())


def ensure_daily_tasks(conn, shop_id, task_date):
    templates_list = conn.execute(
        "SELECT id FROM daily_task_templates WHERE is_active=1 ORDER BY sort_order"
    ).fetchall()
    for t in templates_list:
        conn.execute(
            "INSERT OR IGNORE INTO daily_tasks(shop_id, task_id, task_date) VALUES(?,?,?)",
            (shop_id, t['id'], task_date)
        )
    conn.commit()


def ensure_new_shop_tasks(conn, shop_id):
    templates_list = conn.execute(
        "SELECT id FROM new_shop_task_templates WHERE is_active=1 ORDER BY sort_order"
    ).fetchall()
    for t in templates_list:
        conn.execute(
            "INSERT OR IGNORE INTO new_shop_tasks(shop_id, task_id) VALUES(?,?)",
            (shop_id, t['id'])
        )
    conn.commit()


def migrate_data():
    conn = get_db()
    if os.path.exists(BACKUP_PROFIT):
        try:
            src = sqlite3.connect(BACKUP_PROFIT)
            src.row_factory = sqlite3.Row
            rows = src.execute("SELECT * FROM shops").fetchall()
            for r in rows:
                try:
                    conn.execute("INSERT OR IGNORE INTO shops(name, shop_id, group_name) VALUES(?, '', ?)",
                                 (r['name'], r['group_name'] if 'group_name' in r.keys() else ''))
                except: pass
            rows = src.execute("SELECT * FROM orders").fetchall()
            for r in rows:
                try:
                    conn.execute("""INSERT OR IGNORE INTO orders(douyin_order_no,shop_id,product_name,douyin_amount,
                        taobao_order_no,taobao_amount,refund_status,order_date,buyer_note,system_note,
                        logistics_company,logistics_no,is_influencer,warehouse_status)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (r['douyin_order_no'], r['shop_id'], r['product_name'], r['douyin_amount'],
                         r['taobao_order_no'], r['taobao_amount'], r['refund_status'], r['order_date'],
                         r['buyer_note'], r['system_note'], r['logistics_company'], r['logistics_no'],
                         r['is_influencer'], r['warehouse_status']))
                except: pass
            src.close()
        except: pass

    if os.path.exists(BACKUP_DASHBOARD):
        try:
            src = sqlite3.connect(BACKUP_DASHBOARD)
            src.row_factory = sqlite3.Row
            rows = src.execute("SELECT * FROM shops").fetchall()
            for r in rows:
                try:
                    shop_name = r['shop_name'] if 'shop_name' in r.keys() else r['name']
                    existing = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                    if existing:
                        conn.execute("UPDATE shops SET shop_id=?, group_name=? WHERE id=?",
                                     (r['shop_id'], r['group_name'], existing['id']))
                    else:
                        conn.execute("INSERT INTO shops(name, shop_id, group_name) VALUES(?,?,?)",
                                     (shop_name, r['shop_id'], r['group_name']))
                except: pass
            rows = src.execute("SELECT * FROM daily_snapshot").fetchall()
            for r in rows:
                try:
                    shop_name = None
                    s = src.execute("SELECT shop_name FROM shops WHERE id=?", (r['shop_id'],)).fetchone()
                    if s:
                        shop_name = s['shop_name'] if 'shop_name' in s.keys() else s['name']
                    if shop_name:
                        db_shop = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                        if db_shop:
                            conn.execute("""INSERT OR REPLACE INTO daily_snapshot(
                                shop_id,snapshot_date,status,expire_info,warning,amount,order_count,
                                exposure,clicks,refund_amount,experience_score,pending_ship,
                                pending_after_sale,on_sale,in_warehouse,draft,ship_24h,overdue_ship,
                                after_sale_24h,pending_rectify,last_violation,work_order,fund,
                                experience_fund,insurance_price,valid_orders_30d,cleaned_orders,
                                violation_count,product_score,logistics_score,service_score,
                                abnormal_parcel,recycle_bin,refund_today,qianchuan_id,need_invoice,today_orders,remark
                            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                (db_shop['id'], r['snapshot_date'], r['status'], r['expire_info'],
                                 r['warning'], r['amount'], r['order_count'], r['exposure'], r['clicks'],
                                 r['refund_amount'], r['experience_score'], r['pending_ship'],
                                 r['pending_after_sale'], r['on_sale'], r['in_warehouse'], r['draft'],
                                 r['ship_24h'], r['overdue_ship'], r['after_sale_24h'], r['pending_rectify'],
                                 r['last_violation'], r['work_order'], r['fund'], r['experience_fund'],
                                 r['insurance_price'], r['valid_orders_30d'], r['cleaned_orders'],
                                 r['violation_count'], r['product_score'], r['logistics_score'],
                                 r['service_score'], r['abnormal_parcel'], r['recycle_bin'],
                                 r['refund_today'], r['qianchuan_id'], r['need_invoice'],
                                 r['today_orders'], r['remark']))
                except: pass
            src.close()
        except: pass

    if os.path.exists(BACKUP_TASK):
        try:
            src = sqlite3.connect(BACKUP_TASK)
            src.row_factory = sqlite3.Row
            rows = src.execute("SELECT * FROM shops").fetchall()
            for r in rows:
                try:
                    shop_name = r['shop_name'] if 'shop_name' in r.keys() else r['name']
                    existing = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                    if existing:
                        lid = r['license_id'] if 'license_id' in r.keys() else None
                        st = r['status'] if 'status' in r.keys() else 'normal'
                        nsd = r['new_shop_task_done'] if 'new_shop_task_done' in r.keys() else 0
                        conn.execute("UPDATE shops SET license_id=?, status=?, new_shop_task_done=? WHERE id=?",
                                     (lid, st, nsd, existing['id']))
                    else:
                        lid = r['license_id'] if 'license_id' in r.keys() else None
                        st = r['status'] if 'status' in r.keys() else 'normal'
                        nsd = r['new_shop_task_done'] if 'new_shop_task_done' in r.keys() else 0
                        conn.execute("INSERT INTO shops(name, license_id, status, new_shop_task_done) VALUES(?,?,?,?)",
                                     (shop_name, lid, st, nsd))
                except: pass
            rows = src.execute("SELECT * FROM licenses").fetchall()
            for r in rows:
                try:
                    conn.execute("INSERT OR IGNORE INTO licenses(id,license_name,license_no,holder_name,expire_date,remark) VALUES(?,?,?,?,?,?)",
                                 (r['id'], r['license_name'], r['license_no'], r['holder_name'], r['expire_date'], r['remark']))
                except: pass
            rows = src.execute("SELECT * FROM platform_accounts").fetchall()
            for r in rows:
                try:
                    shop_name = None
                    s = src.execute("SELECT shop_name FROM shops WHERE id=?", (r['shop_id'],)).fetchone()
                    if s:
                        shop_name = s['shop_name'] if 'shop_name' in s.keys() else s['name']
                    if shop_name:
                        db_shop = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                        if db_shop:
                            conn.execute("INSERT OR IGNORE INTO platform_accounts(shop_id,platform_name,account,password_enc,remark) VALUES(?,?,?,?,?)",
                                         (db_shop['id'], r['platform_name'], r['account'], r['password_enc'], r['remark']))
                except: pass
            rows = src.execute("SELECT * FROM daily_task_templates").fetchall()
            for r in rows:
                try:
                    conn.execute("INSERT OR IGNORE INTO daily_task_templates(id,task_name,sort_order,is_active) VALUES(?,?,?,?)",
                                 (r['id'], r['task_name'], r['sort_order'], r['is_active']))
                except: pass
            rows = src.execute("SELECT * FROM new_shop_task_templates").fetchall()
            for r in rows:
                try:
                    conn.execute("INSERT OR IGNORE INTO new_shop_task_templates(id,task_name,sort_order,is_active) VALUES(?,?,?,?)",
                                 (r['id'], r['task_name'], r['sort_order'], r['is_active']))
                except: pass
            rows = src.execute("SELECT * FROM daily_tasks").fetchall()
            for r in rows:
                try:
                    shop_name = None
                    s = src.execute("SELECT shop_name FROM shops WHERE id=?", (r['shop_id'],)).fetchone()
                    if s:
                        shop_name = s['shop_name'] if 'shop_name' in s.keys() else s['name']
                    if shop_name:
                        db_shop = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                        if db_shop:
                            conn.execute("INSERT OR IGNORE INTO daily_tasks(shop_id,task_id,task_date,is_completed,completed_at,remark) VALUES(?,?,?,?,?,?)",
                                         (db_shop['id'], r['task_id'], r['task_date'], r['is_completed'], r['completed_at'], r['remark']))
                except: pass
            rows = src.execute("SELECT * FROM new_shop_tasks").fetchall()
            for r in rows:
                try:
                    shop_name = None
                    s = src.execute("SELECT shop_name FROM shops WHERE id=?", (r['shop_id'],)).fetchone()
                    if s:
                        shop_name = s['shop_name'] if 'shop_name' in s.keys() else s['name']
                    if shop_name:
                        db_shop = conn.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
                        if db_shop:
                            conn.execute("INSERT OR IGNORE INTO new_shop_tasks(shop_id,task_id,is_completed,completed_at,remark) VALUES(?,?,?,?,?)",
                                         (db_shop['id'], r['task_id'], r['is_completed'], r['completed_at'], r['remark']))
                except: pass
            src.close()
        except: pass

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS shops (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            shop_id TEXT DEFAULT '',
            group_name TEXT DEFAULT '',
            status TEXT DEFAULT 'normal',
            license_id INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            new_shop_task_done INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS orders (
            douyin_order_no TEXT PRIMARY KEY,
            shop_id INTEGER NOT NULL,
            product_name TEXT NOT NULL,
            douyin_amount REAL NOT NULL,
            taobao_order_no TEXT DEFAULT '',
            taobao_amount REAL DEFAULT 0,
            refund_status TEXT DEFAULT '待发货',
            order_date TEXT NOT NULL,
            buyer_note TEXT DEFAULT '',
            system_note TEXT DEFAULT '',
            logistics_company TEXT DEFAULT '',
            logistics_no TEXT DEFAULT '',
            is_influencer TEXT DEFAULT '否',
            warehouse_status TEXT DEFAULT '未到仓库'
        );
        CREATE TABLE IF NOT EXISTS daily_snapshot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            status TEXT DEFAULT '',
            expire_info TEXT DEFAULT '',
            warning INTEGER DEFAULT 0,
            amount REAL DEFAULT 0,
            order_count INTEGER DEFAULT 0,
            exposure INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            refund_amount REAL DEFAULT 0,
            experience_score TEXT DEFAULT '',
            pending_ship INTEGER DEFAULT 0,
            pending_after_sale INTEGER DEFAULT 0,
            on_sale INTEGER DEFAULT 0,
            in_warehouse INTEGER DEFAULT 0,
            draft INTEGER DEFAULT 0,
            ship_24h TEXT DEFAULT '',
            overdue_ship INTEGER DEFAULT 0,
            after_sale_24h TEXT DEFAULT '',
            pending_rectify INTEGER DEFAULT 0,
            last_violation TEXT DEFAULT '',
            work_order INTEGER DEFAULT 0,
            fund REAL DEFAULT 0,
            experience_fund REAL DEFAULT 0,
            insurance_price TEXT DEFAULT '',
            valid_orders_30d INTEGER DEFAULT 0,
            cleaned_orders INTEGER DEFAULT 0,
            violation_count INTEGER DEFAULT 0,
            product_score TEXT DEFAULT '',
            logistics_score TEXT DEFAULT '',
            service_score TEXT DEFAULT '',
            abnormal_parcel INTEGER DEFAULT 0,
            recycle_bin INTEGER DEFAULT 0,
            refund_today REAL DEFAULT 0,
            qianchuan_id TEXT DEFAULT '',
            need_invoice REAL DEFAULT 0,
            today_orders INTEGER DEFAULT 0,
            remark TEXT DEFAULT '',
            UNIQUE(shop_id, snapshot_date)
        );
        CREATE TABLE IF NOT EXISTS licenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            license_name TEXT NOT NULL,
            license_no TEXT DEFAULT '',
            holder_name TEXT DEFAULT '',
            expire_date TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS platform_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL,
            platform_name TEXT NOT NULL,
            account TEXT DEFAULT '',
            password_enc TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (shop_id) REFERENCES shops(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS daily_task_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_name TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS new_shop_task_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_name TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS daily_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL,
            task_id INTEGER NOT NULL,
            task_date TEXT NOT NULL,
            is_completed INTEGER DEFAULT 0,
            completed_at TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            UNIQUE(shop_id, task_id, task_date),
            FOREIGN KEY (shop_id) REFERENCES shops(id) ON DELETE CASCADE,
            FOREIGN KEY (task_id) REFERENCES daily_task_templates(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS new_shop_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL,
            task_id INTEGER NOT NULL,
            is_completed INTEGER DEFAULT 0,
            completed_at TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            UNIQUE(shop_id, task_id),
            FOREIGN KEY (shop_id) REFERENCES shops(id) ON DELETE CASCADE,
            FOREIGN KEY (task_id) REFERENCES new_shop_task_templates(id) ON DELETE CASCADE
        );
    """)

    for col, typ in [("shop_id", "TEXT DEFAULT ''"), ("status", "TEXT DEFAULT 'normal"),
                     ("license_id", "INTEGER DEFAULT NULL"), ("new_shop_task_done", "INTEGER DEFAULT 0")]:
        try: conn.execute(f"ALTER TABLE shops ADD COLUMN {col} {typ}")
        except: pass

    daily_count = conn.execute("SELECT COUNT(*) FROM daily_task_templates").fetchone()[0]
    if daily_count == 0:
        defaults = [
            ("处理待发货订单", 1), ("处理售后/退款订单", 2),
            ("检查店铺违规", 3), ("检查体验分并优化", 4),
            ("上新商品（3-5个）", 5), ("优化商品标题/主图", 6),
            ("检查库存情况", 7), ("处理工单", 8),
            ("店铺数据复盘", 9), ("竞品分析", 10),
        ]
        conn.executemany("INSERT INTO daily_task_templates(task_name, sort_order) VALUES(?,?)", defaults)

    new_count = conn.execute("SELECT COUNT(*) FROM new_shop_task_templates").fetchone()[0]
    if new_count == 0:
        defaults = [
            ("完善店铺基本信息（名称、头像、简介）", 1),
            ("设置运费模板", 2), ("设置售后地址", 3),
            ("开通支付方式", 4), ("上传营业执照（如需要）", 5),
            ("完善店铺资质", 6), ("设置客服自动回复", 7),
            ("完善退换货规则", 8), ("上架第一批商品（5-10个）", 9),
            ("设置店铺优惠券", 10), ("开通运费险", 11),
            ("店铺装修", 12),
        ]
        conn.executemany("INSERT INTO new_shop_task_templates(task_name, sort_order) VALUES(?,?)", defaults)

    conn.commit()
    conn.close()

    migrate_data()
    auto_import_all()


@app.get("/", response_class=HTMLResponse)
def root_redirect(request: Request):
    return RedirectResponse("/profit/", status_code=302)


# ==================== Profit Tracker Routes ====================

@profit_router.get("/", response_class=HTMLResponse)
def profit_index(request: Request):
    db = get_db()
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    month_start = date.today().replace(day=1).isoformat()

    def sf(d):
        return dict(db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(douyin_amount),0) as dy, "
            "COALESCE(SUM(taobao_amount),0) as tb, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
            "ELSE 6.9 END),0) as freight, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END),0) as commission, "
            "COALESCE(SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END),0) as ship_cost, "
            "COALESCE(SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
            "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END),0) as return_cost "
            "FROM orders WHERE order_date=?", (d,)
        ).fetchone())

    def sf2(a, b):
        return dict(db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(douyin_amount),0) as dy, "
            "COALESCE(SUM(taobao_amount),0) as tb, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
            "ELSE 6.9 END),0) as freight, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END),0) as commission, "
            "COALESCE(SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END),0) as ship_cost, "
            "COALESCE(SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
            "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END),0) as return_cost "
            "FROM orders WHERE order_date>=? AND order_date<=?", (a, b)
        ).fetchone())

    td = sf(today); td["profit"] = td["dy"] - td["tb"] - td["freight"] - td["commission"] - td["ship_cost"] - td["return_cost"]
    wk = sf2(week_ago, today); wk["profit"] = wk["dy"] - wk["tb"] - wk["freight"] - wk["commission"] - wk["ship_cost"] - wk["return_cost"]
    mn = sf2(month_start, today); mn["profit"] = mn["dy"] - mn["tb"] - mn["freight"] - mn["commission"] - mn["ship_cost"] - mn["return_cost"]

    trend = [dict(r) for r in db.execute(
        "SELECT order_date, "
        "SUM(douyin_amount)-SUM(taobao_amount)-"
        "SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
        "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
        "ELSE 6.9 END)-"
        "SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END)-"
        "SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        "SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        "FROM orders WHERE order_date>=date('now','-7 days') GROUP BY order_date ORDER BY order_date"
    ).fetchall()]

    sp = [dict(r) for r in db.execute(
        "SELECT s.name, s.id, s.group_name, "
        "SUM(o.douyin_amount)-SUM(o.taobao_amount)-"
        "SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        "WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        "WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        "ELSE 6.9 END)-"
        "SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END)-"
        "SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        "SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        "WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        "FROM orders o JOIN shops s ON o.shop_id=s.id WHERE o.order_date>=date('now','-30 days') "
        "GROUP BY o.shop_id ORDER BY profit DESC"
    ).fetchall()]

    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]

    db.close()
    return templates.TemplateResponse(request, "profit/index.html", {"td": td, "wk": wk, "mn": mn, "trend": trend, "shop_profit": sp, "groups": [g["group_name"] for g in groups]})


@profit_router.get("/api/trend")
def profit_api_trend(date_from: Optional[str] = None, date_to: Optional[str] = None, shop_id: Optional[int] = None, group: Optional[str] = None):
    db = get_db()
    df = date_from or date.today().isoformat()
    dt = date_to or date.today().isoformat()
    c, p = ["o.order_date>=?", "o.order_date<=?"], [df, dt]
    if shop_id: c.append("o.shop_id=?"); p.append(shop_id)
    if group: c.append("s.group_name=?"); p.append(group)
    w = " AND ".join(c)
    trend = [dict(r) for r in db.execute(
        f"SELECT o.order_date, "
        f"SUM(o.douyin_amount)-SUM(o.taobao_amount)-"
        f"SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        f"ELSE 6.9 END)-"
        f"SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END)-"
        f"SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        f"SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        f"WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        f"FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE {w} GROUP BY o.order_date ORDER BY o.order_date", p
    ).fetchall()]
    db.close()
    return trend


@profit_router.get("/orders", response_class=HTMLResponse)
def profit_orders_page(request: Request, shop_id: str = "", date_from: str = "",
                date_to: str = "", status: str = "",
                keyword: str = "", page: int = 1, group: str = ""):
    db = get_db()
    pp = 20
    c, p = [], []
    if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
    if date_from: c.append("o.order_date>=?"); p.append(date_from)
    if date_to: c.append("o.order_date<=?"); p.append(date_to)
    if status: c.append("o.refund_status=?"); p.append(status)
    if keyword: c.append("(o.douyin_order_no LIKE ? OR o.product_name LIKE ? OR o.taobao_order_no LIKE ?)"); lk = f"%{keyword}%"; p.extend([lk,lk,lk])
    if group: c.append("s.group_name=?"); p.append(group)
    w = ("WHERE " + " AND ".join(c)) if c else ""
    total = db.execute(f"SELECT COUNT(*) FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w}", p).fetchone()[0]
    items = [dict(r) for r in db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w} ORDER BY o.order_date DESC, o.douyin_order_no DESC LIMIT ? OFFSET ?", p + [pp, (page-1)*pp]).fetchall()]
    shops = [dict(r) for r in db.execute("SELECT * FROM shops ORDER BY id").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    today = date.today().isoformat()
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    d7 = (date.today() - timedelta(days=6)).isoformat()
    d15 = (date.today() - timedelta(days=14)).isoformat()
    d30 = (date.today() - timedelta(days=29)).isoformat()
    return templates.TemplateResponse(request, "profit/orders.html", {"items": items, "shops": shops, "total": total, "page": page, "per_page": pp, "f_sid": shop_id, "f_df": date_from, "f_dt": date_to, "f_st": status, "f_kw": keyword, "f_group": group, "today": today, "yesterday": yesterday, "d7": d7, "d15": d15, "d30": d30, "groups": [g["group_name"] for g in groups]})


@profit_router.get("/orders/new", response_class=HTMLResponse)
def profit_order_new(request: Request):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "profit/order_form.html", {"order": None, "shops": shops})


@profit_router.get("/orders/export")
def profit_export_orders(shop_id: str = "", date_from: str = "",
                  date_to: str = "", status: str = "",
                  keyword: str = "", selected: str = "", group: str = ""):
    db = get_db()
    if selected:
        order_list = selected.split(",")
        placeholders = ",".join(["?" for _ in order_list])
        rows = db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no IN ({placeholders}) ORDER BY o.order_date DESC", order_list).fetchall()
    else:
        c, p = [], []
        if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
        if date_from: c.append("o.order_date>=?"); p.append(date_from)
        if date_to: c.append("o.order_date<=?"); p.append(date_to)
        if status: c.append("o.refund_status=?"); p.append(status)
        if keyword: c.append("(o.douyin_order_no LIKE ? OR o.product_name LIKE ? OR o.taobao_order_no LIKE ?)"); lk = f"%{keyword}%"; p.extend([lk,lk,lk])
        if group: c.append("s.group_name=?"); p.append(group)
        w = ("WHERE " + " AND ".join(c)) if c else ""
        rows = db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w} ORDER BY o.order_date DESC", p).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单数据"
    ws.append(["店铺", "抖音订单号", "商品名称", "抖音金额", "淘宝订单号", "淘宝金额", "利润", "订单状态", "订单日期", "买家备注", "系统备注", "物流公司", "物流单号", "达人带货", "仓库状态", "运费", "达人佣金", "发货成本", "退货成本"])
    for r in rows:
        freight = 6.9
        commission = r["douyin_amount"] * 0.25 if r["is_influencer"] == "是" and r["refund_status"] in ("待发货", "已发货") else 0
        ship_cost = 5.5 if r["warehouse_status"] in ("已到达仓库未发货", "已到仓库已发货") and r["refund_status"] in ("待发货", "已发货") else 0
        return_cost = 3.5 if (r["warehouse_status"] == "已到达仓库未发货" and r["refund_status"] not in ("待发货", "已发货")) or r["refund_status"] in ("已发货退款", "退货退款") else 0
        if r["refund_status"] in ("已发货退款", "退货退款"):
            freight = 6.9
        elif r["refund_status"] not in ("待发货",) and r["warehouse_status"] == "未到仓库":
            freight = 0
        elif r["refund_status"] not in ("待发货",) and r["warehouse_status"] == "已到达仓库未发货":
            freight = 0
        else:
            freight = 6.9
        profit = r["douyin_amount"] - r["taobao_amount"] - freight - commission - ship_cost - return_cost
        ws.append([r["shop_name"], r["douyin_order_no"], r["product_name"], r["douyin_amount"],
                   r["taobao_order_no"], r["taobao_amount"], profit, r["refund_status"], r["order_date"],
                   r["buyer_note"], r["system_note"], r["logistics_company"], r["logistics_no"],
                   r["is_influencer"], r["warehouse_status"], freight, commission, ship_cost, return_cost])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=orders_export.xlsx"})


@profit_router.get("/orders/{dy_no}", response_class=HTMLResponse)
def profit_order_detail(request: Request, dy_no: str):
    db = get_db()
    order = db.execute("SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no=?", (dy_no,)).fetchone()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    if not order:
        return RedirectResponse("/profit/orders", status_code=302)
    return templates.TemplateResponse(request, "profit/order_detail.html", {"order": dict(order), "shops": shops})


@profit_router.get("/api/orders/{dy_no}/json")
def profit_api_order_json(dy_no: str):
    db = get_db()
    order = db.execute("SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no=?", (dy_no,)).fetchone()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    if not order:
        return {"error": "not found"}
    return {"order": dict(order), "shops": shops}


@profit_router.get("/api/orders/by-date/{date_str}")
def profit_api_orders_by_date(date_str: str):
    db = get_db()
    orders = [dict(r) for r in db.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.order_date=? ORDER BY o.douyin_order_no", (date_str,)
    ).fetchall()]
    for o in orders:
        o["profit"] = calc_profit(o["douyin_amount"], o["taobao_amount"], o["refund_status"], o["is_influencer"], o["warehouse_status"])
    db.close()
    return {"orders": orders}


@profit_router.get("/api/shop/{sid}/orders")
def profit_api_shop_orders(sid: int):
    db = get_db()
    shop = db.execute("SELECT * FROM shops WHERE id=?", (sid,)).fetchone()
    orders = [dict(r) for r in db.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.shop_id=? ORDER BY o.order_date DESC", (sid,)
    ).fetchall()]
    for o in orders:
        o["profit"] = calc_profit(o["douyin_amount"], o["taobao_amount"], o["refund_status"], o["is_influencer"], o["warehouse_status"])
    db.close()
    return {"shop": dict(shop) if shop else None, "orders": orders}


@profit_router.get("/import", response_class=HTMLResponse)
def profit_import_page(request: Request, msg: Optional[str] = None):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "profit/import.html", {"shops": shops, "msg": msg or ""})


@profit_router.get("/shops", response_class=HTMLResponse)
def profit_shops_page(request: Request, msg: Optional[str] = None):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops ORDER BY id").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "profit/shops.html", {"shops": shops, "msg": msg or "", "groups": [g["group_name"] for g in groups]})


@profit_router.post("/api/shops/add")
def profit_api_add_shop(name: str = Form(...), group_name: str = Form("")):
    db = get_db()
    try:
        db.execute("INSERT INTO shops(name, group_name) VALUES(?, ?)", (name, group_name))
        db.commit()
    except sqlite3.IntegrityError:
        pass
    db.close()
    return RedirectResponse("/profit/shops?msg=店铺添加成功", status_code=302)


@profit_router.post("/api/shops/{sid}/update")
def profit_api_update_shop(sid: int, name: str = Form(...), group_name: str = Form("")):
    db = get_db()
    db.execute("UPDATE shops SET name=?, group_name=? WHERE id=?", (name, group_name, sid))
    db.commit()
    db.close()
    return RedirectResponse("/profit/shops", status_code=302)


@profit_router.get("/api/groups")
def profit_api_get_groups():
    db = get_db()
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return [g["group_name"] for g in groups]


@profit_router.post("/api/shops/{sid}/delete")
def profit_api_delete_shop(sid: int):
    db = get_db()
    db.execute("DELETE FROM shops WHERE id=?", (sid,))
    db.commit()
    db.close()
    return RedirectResponse("/profit/shops?msg=店铺已删除", status_code=302)


@profit_router.get("/stats", response_class=HTMLResponse)
def profit_stats_page(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None, shop_id: str = "", days: Optional[int] = None, group: Optional[str] = None):
    if days:
        df = (date.today() - timedelta(days=days-1)).isoformat()
    else:
        df = date_from or (date.today() - timedelta(days=29)).isoformat()
    dt = date_to or date.today().isoformat()
    db = get_db()
    c, p = ["o.order_date>=?", "o.order_date<=?"], [df, dt]
    if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
    if group: c.append("s.group_name=?"); p.append(group)
    w = " AND ".join(c)
    daily = [dict(r) for r in db.execute(
        f"SELECT o.order_date, COUNT(*) as order_count, SUM(o.douyin_amount) as total_douyin, SUM(o.taobao_amount) as total_taobao, "
        f"SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        f"ELSE 6.9 END) as total_freight, "
        f"SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END) as total_commission, "
        f"SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END) as total_ship, "
        f"SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        f"WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as total_return "
        f"FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE {w} GROUP BY o.order_date ORDER BY o.order_date", p
    ).fetchall()]
    for row in daily: row["profit"] = row["total_douyin"] - row["total_taobao"] - row["total_freight"] - row["total_commission"] - row["total_ship"] - row["total_return"]
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "profit/stats.html", {"daily": daily, "shops": shops, "f_df": df, "f_dt": dt, "f_sid": shop_id, "f_days": days or "", "f_group": group or "", "groups": [g["group_name"] for g in groups]})


@profit_router.post("/api/orders/import")
async def profit_api_import(file: UploadFile = File(...), shop_id: int = Form(None)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    cm = {}
    for i, h in enumerate(headers):
        if h == "订单号": cm["dy_no"] = i
        elif h == "店铺名称": cm["shop"] = i
        elif h == "下单时间": cm["date"] = i
        elif h == "订单状态": cm["status"] = i
        elif h == "买家备注": cm["buyer_note"] = i
        elif h == "系统备注": cm["sys_note"] = i
        elif h == "订单金额": cm["dy_amt"] = i
        elif h == "物流公司名称": cm["logistics"] = i
        elif h == "物流单号": cm["logistics_no"] = i
        elif h == "商品名称": cm["name"] = i
        elif h == "子订单状态": cm["sub_status"] = i
    db = get_db()
    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        def g(key, default=None):
            i = cm.get(key)
            return row[i] if i is not None and i < len(row) and row[i] is not None else default
        dy_no = str(g("dy_no", ""))
        name = str(g("name", ""))
        if not dy_no and not name: continue
        dy_amt = float(g("dy_amt", 0) or 0)
        order_date = str(g("date", date.today().isoformat()))[:10]
        status = str(g("status", ""))
        sub_status = str(g("sub_status", ""))
        if "退" in str(status) or "退" in str(sub_status):
            rf = "退货退款"
        elif "发" in str(status) or "发" in str(sub_status):
            rf = "已发货"
        else:
            rf = "待发货"
        real_shop_id = shop_id
        shop_name = g("shop")
        if shop_name:
            shop_name = str(shop_name).strip()
            row_s = db.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
            if row_s:
                real_shop_id = row_s["id"]
            else:
                cur = db.execute("INSERT INTO shops(name, group_name) VALUES(?, '')", (shop_name,))
                real_shop_id = cur.lastrowid
        if not real_shop_id:
            default_shop = db.execute("SELECT id FROM shops WHERE name='默认店铺'").fetchone()
            if default_shop:
                real_shop_id = default_shop["id"]
            else:
                cur = db.execute("INSERT INTO shops(name, group_name) VALUES('默认店铺', '')")
                real_shop_id = cur.lastrowid
        try:
            db.execute("INSERT INTO orders(douyin_order_no,shop_id,product_name,douyin_amount,taobao_order_no,taobao_amount,refund_status,order_date,buyer_note,system_note,logistics_company,logistics_no,is_influencer,warehouse_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                       (dy_no, real_shop_id, name, dy_amt, "", 0, rf, order_date,
                        str(g("buyer_note", "")), str(g("sys_note", "")), str(g("logistics", "")), str(g("logistics_no", "")), "否", "未到仓库"))
            cnt += 1
        except sqlite3.IntegrityError:
            pass
    db.commit(); db.close(); wb.close()
    return RedirectResponse(f"/profit/import?msg=成功导入{cnt}条订单", status_code=302)


@profit_router.post("/api/orders")
def profit_api_create(shop_id: int = Form(...), douyin_order_no: str = Form(...), product_name: str = Form(...),
               douyin_amount: float = Form(...), taobao_order_no: str = Form(""), taobao_amount: float = Form(0),
               refund_status: str = Form("待发货"), order_date: str = Form(...),
               buyer_note: str = Form(""), system_note: str = Form(""), logistics_company: str = Form(""), logistics_no: str = Form(""),
               is_influencer: str = Form("否"), warehouse_status: str = Form("未到仓库")):
    db = get_db()
    try:
        db.execute("INSERT INTO orders(douyin_order_no,shop_id,product_name,douyin_amount,taobao_order_no,taobao_amount,refund_status,order_date,buyer_note,system_note,logistics_company,logistics_no,is_influencer,warehouse_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                   (douyin_order_no, shop_id, product_name, douyin_amount, taobao_order_no, taobao_amount, refund_status, order_date, buyer_note, system_note, logistics_company, logistics_no, is_influencer, warehouse_status))
        db.commit()
    except sqlite3.IntegrityError:
        pass
    db.close()
    return RedirectResponse("/profit/orders", status_code=302)


@profit_router.post("/api/orders/{dy_no}")
def profit_api_update(dy_no: str, shop_id: int = Form(...), product_name: str = Form(...),
               douyin_amount: float = Form(...), taobao_order_no: str = Form(""), taobao_amount: float = Form(0),
               refund_status: str = Form("待发货"), order_date: str = Form(...),
               buyer_note: str = Form(""), system_note: str = Form(""), logistics_company: str = Form(""), logistics_no: str = Form(""),
               is_influencer: str = Form("否"), warehouse_status: str = Form("未到仓库")):
    db = get_db()
    db.execute("UPDATE orders SET shop_id=?,product_name=?,douyin_amount=?,taobao_order_no=?,taobao_amount=?,refund_status=?,order_date=?,buyer_note=?,system_note=?,logistics_company=?,logistics_no=?,is_influencer=?,warehouse_status=? WHERE douyin_order_no=?",
               (shop_id, product_name, douyin_amount, taobao_order_no, taobao_amount, refund_status, order_date, buyer_note, system_note, logistics_company, logistics_no, is_influencer, warehouse_status, dy_no))
    db.commit(); db.close()
    return RedirectResponse("/profit/orders", status_code=302)


@profit_router.post("/api/orders/{dy_no}/delete")
def profit_api_delete(dy_no: str):
    db = get_db()
    db.execute("DELETE FROM orders WHERE douyin_order_no=?", (dy_no,))
    db.commit(); db.close()
    return RedirectResponse("/profit/orders", status_code=302)


@profit_router.post("/api/shops")
def profit_api_create_shop(name: str = Form(...)):
    db = get_db()
    try: db.execute("INSERT INTO shops(name) VALUES(?)", (name,)); db.commit()
    except: pass
    db.close()
    return RedirectResponse("/profit", status_code=302)


@profit_router.get("/api/template")
def profit_download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单号", "店铺名称", "下单时间", "订单状态", "买家备注", "系统备注", "订单金额", "物流公司名称", "物流单号", "商品名称", "子订单状态"])
    ws.append(["DY20240101001", "默认店铺", "2024-01-01 12:00:00", "已付款", "", "", 99.99, "顺丰速运", "SF1234567890", "示例商品", "已付款"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template.xlsx"})


# ==================== Dashboard Routes ====================

@dashboard_router.get("/", response_class=HTMLResponse)
def dashboard_index(request: Request, qdate: str = ""):
    conn = get_db()
    today = date.today().isoformat()
    auto_import_all()

    available_dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC LIMIT 60"
    ).fetchall()]

    if not available_dates:
        conn.close()
        return templates.TemplateResponse(request, "dashboard/index.html", {
            "request": request, "summary": None, "shops": [], "today": today,
            "available_dates": [], "selected_date": today
        })

    selected_date = qdate if qdate in available_dates else available_dates[0]
    summary = get_summary(conn, selected_date)

    sd = date.fromisoformat(selected_date)
    summary_7d = get_range_summary(conn, (sd - timedelta(days=6)).isoformat(), selected_date)
    summary_15d = get_range_summary(conn, (sd - timedelta(days=14)).isoformat(), selected_date)
    summary_30d = get_range_summary(conn, (sd - timedelta(days=29)).isoformat(), selected_date)

    shops = [dict(r) for r in conn.execute("""
        SELECT s.name as shop_name, s.group_name, d.* FROM daily_snapshot d
        JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date=?
        ORDER BY d.amount DESC
    """, (selected_date,)).fetchall()]

    groups = sorted(set(s['group_name'] for s in shops if s['group_name']))

    conn.close()
    return templates.TemplateResponse(request, "dashboard/index.html", {
        "request": request, "summary": summary, "shops": shops,
        "today": today, "available_dates": available_dates, "selected_date": selected_date,
        "groups": groups, "summary_7d": summary_7d, "summary_15d": summary_15d, "summary_30d": summary_30d
    })


@dashboard_router.get("/api/summary/range_detail")
def dashboard_api_summary_range_detail(start: str, end: str, metric: str):
    conn = get_db()
    metric_map = {
        'total_amount': 'SUM(d.amount)',
        'total_orders': 'SUM(d.order_count)',
        'total_today_orders': 'SUM(d.order_count)',
        'total_refund': 'SUM(d.refund_amount)',
        'total_refund_today': 'SUM(d.refund_today)',
        'total_pending_ship': 'SUM(d.pending_ship)',
        'total_overdue_ship': 'SUM(d.overdue_ship)',
        'total_exposure': 'SUM(d.exposure)',
        'total_clicks': 'SUM(d.clicks)',
    }
    col = metric_map.get(metric, 'SUM(d.amount)')
    shops = [dict(r) for r in conn.execute(f"""
        SELECT s.name as shop_name, {col} as val
        FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id
        WHERE d.snapshot_date >= ? AND d.snapshot_date <= ?
        GROUP BY d.shop_id HAVING val > 0 ORDER BY val DESC
    """, (start, end)).fetchall()]
    conn.close()
    return {"shops": shops, "metric": metric, "start": start, "end": end}


@dashboard_router.get("/ranking", response_class=HTMLResponse)
def dashboard_ranking_page(request: Request, sort_by: str = "amount", days: int = 1, group: str = ""):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    valid_sorts = {
        "amount": "SUM(d.amount) DESC",
        "orders": "SUM(d.order_count) DESC",
        "refund": "SUM(d.refund_amount) DESC",
        "exposure": "SUM(d.exposure) DESC",
        "clicks": "SUM(d.clicks) DESC",
    }
    order_clause = valid_sorts.get(sort_by, "SUM(d.amount) DESC")

    group_filter = ""
    params = [start_date, today]
    if group:
        group_filter = " AND s.group_name = ?"
        params.append(group)

    shops = [dict(r) for r in conn.execute(f"""
        SELECT s.name as shop_name, s.group_name,
               COUNT(*) as days_count,
               SUM(d.amount) as total_amount,
               SUM(d.order_count) as total_orders,
               SUM(d.refund_amount) as total_refund,
               SUM(d.exposure) as total_exposure,
               SUM(d.clicks) as total_clicks,
               SUM(d.today_orders) as total_today_orders,
               SUM(d.pending_ship) as total_pending_ship,
               SUM(d.overdue_ship) as total_overdue_ship
        FROM daily_snapshot d JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date >= ? AND d.snapshot_date <= ? {group_filter}
        GROUP BY d.shop_id
        ORDER BY {order_clause}
    """, params).fetchall()]

    all_groups = sorted(set(r['group_name'] for r in conn.execute(
        "SELECT DISTINCT s.group_name FROM shops s JOIN daily_snapshot d ON d.shop_id=s.id WHERE s.group_name != ''"
    ).fetchall()))

    conn.close()
    return templates.TemplateResponse(request, "dashboard/ranking.html", {
        "request": request, "shops": shops, "sort_by": sort_by,
        "days": days, "today": today, "group": group, "all_groups": all_groups
    })


@dashboard_router.get("/trends", response_class=HTMLResponse)
def dashboard_trends_page(request: Request, days: int = 7):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    daily = [dict(r) for r in conn.execute("""
        SELECT snapshot_date,
               SUM(amount) as total_amount,
               SUM(order_count) as total_orders,
               SUM(refund_amount) as total_refund,
               SUM(exposure) as total_exposure,
               SUM(clicks) as total_clicks,
               SUM(pending_ship) as total_pending_ship,
               SUM(overdue_ship) as total_overdue_ship
        FROM daily_snapshot
        WHERE snapshot_date >= ? AND snapshot_date <= ?
        GROUP BY snapshot_date ORDER BY snapshot_date
    """, (start_date, today)).fetchall()]

    all_shops = [dict(r) for r in conn.execute("SELECT DISTINCT s.name as shop_name FROM shops s JOIN daily_snapshot d ON d.shop_id=s.id ORDER BY s.name").fetchall()]

    conn.close()
    return templates.TemplateResponse(request, "dashboard/trends.html", {
        "request": request, "daily": daily, "days": days, "today": today,
        "all_shops": all_shops
    })


@dashboard_router.get("/alerts", response_class=HTMLResponse)
def dashboard_alerts_page(request: Request):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    available_dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC LIMIT 1"
    ).fetchall()]

    if not available_dates:
        conn.close()
        return templates.TemplateResponse(request, "dashboard/alerts.html", {
            "request": request, "alerts": [], "today": today
        })

    selected_date = available_dates[0]

    shops = [dict(r) for r in conn.execute("""
        SELECT s.name as shop_name, s.group_name, d.* FROM daily_snapshot d
        JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date=?
    """, (selected_date,)).fetchall()]

    alerts = []
    for shop in shops:
        reasons = []
        if shop['overdue_ship'] > 0:
            reasons.append(f"超时未发货 {shop['overdue_ship']} 单")
        exp = safe_float(shop['experience_score'])
        if exp > 0 and exp < 70:
            reasons.append(f"体验分偏低: {shop['experience_score']}")
        if shop['violation_count'] > 0:
            reasons.append(f"违规 {shop['violation_count']} 次")
        if shop['pending_after_sale'] > 0:
            reasons.append(f"待售后 {shop['pending_after_sale']} 单")
        if shop['abnormal_parcel'] > 0:
            reasons.append(f"异常包裹 {shop['abnormal_parcel']} 个")
        if safe_float(shop['last_violation']):
            reasons.append(f"最近违规: {shop['last_violation']}")
        if reasons:
            alerts.append({"shop": shop, "reasons": reasons})

    alerts.sort(key=lambda x: len(x['reasons']), reverse=True)
    conn.close()
    return templates.TemplateResponse(request, "dashboard/alerts.html", {
        "request": request, "alerts": alerts, "today": today
    })


@dashboard_router.get("/api/summary/detail")
def dashboard_api_summary_detail(date: str, metric: str):
    conn = get_db()
    shops = []
    if metric == 'total_orders':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.order_count as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.order_count>0 ORDER BY d.order_count DESC", (date,)).fetchall()]
    elif metric == 'total_today_orders':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.today_orders as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.today_orders>0 ORDER BY d.today_orders DESC", (date,)).fetchall()]
    elif metric == 'total_amount':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.amount as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.amount>0 ORDER BY d.amount DESC", (date,)).fetchall()]
    elif metric == 'total_refund':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.refund_amount as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.refund_amount>0 ORDER BY d.refund_amount DESC", (date,)).fetchall()]
    elif metric == 'total_refund_today':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.refund_today as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.refund_today>0 ORDER BY d.refund_today DESC", (date,)).fetchall()]
    elif metric == 'total_pending_ship':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.pending_ship as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.pending_ship>0 ORDER BY d.pending_ship DESC", (date,)).fetchall()]
    elif metric == 'total_overdue_ship':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.overdue_ship as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.overdue_ship>0 ORDER BY d.overdue_ship DESC", (date,)).fetchall()]
    elif metric == 'total_exposure':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.exposure as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? ORDER BY d.exposure DESC", (date,)).fetchall()]
    elif metric == 'total_clicks':
        shops = [dict(r) for r in conn.execute("SELECT s.name as shop_name, d.clicks as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? ORDER BY d.clicks DESC", (date,)).fetchall()]
    conn.close()
    return {"shops": shops, "metric": metric, "date": date}


@dashboard_router.get("/api/shop/detail")
def dashboard_api_shop_detail(shop_name: str, range: str = "1"):
    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE name=?", (shop_name,)).fetchone()
    if not shop:
        conn.close()
        return {"error": "not found"}
    days = int(range) if range.isdigit() else 1
    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()
    history = [dict(r) for r in conn.execute("""
        SELECT * FROM daily_snapshot WHERE shop_id=? AND snapshot_date>=? AND snapshot_date<=?
        ORDER BY snapshot_date
    """, (shop['id'], start_date, today)).fetchall()]
    conn.close()
    return {"shop": dict(shop), "history": history, "days": days}


@dashboard_router.get("/api/dates")
def dashboard_api_dates():
    conn = get_db()
    dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date"
    ).fetchall()]
    conn.close()
    return {"dates": dates}


@dashboard_router.get("/shop/{shop_name}", response_class=HTMLResponse)
def dashboard_shop_detail(request: Request, shop_name: str, days: int = 30):
    conn = get_db()
    auto_import_all()

    shop = conn.execute("SELECT * FROM shops WHERE name=?", (shop_name,)).fetchone()
    if not shop:
        conn.close()
        return RedirectResponse("/dashboard", status_code=302)

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    history = [dict(r) for r in conn.execute("""
        SELECT * FROM daily_snapshot
        WHERE shop_id=? AND snapshot_date >= ? AND snapshot_date <= ?
        ORDER BY snapshot_date
    """, (shop['id'], start_date, today)).fetchall()]

    latest = history[-1] if history else None
    conn.close()
    return templates.TemplateResponse(request, "dashboard/shop_detail.html", {
        "request": request, "shop": dict(shop), "history": history,
        "latest": latest, "days": days, "today": today
    })


@dashboard_router.get("/import", response_class=HTMLResponse)
def dashboard_import_page(request: Request, msg: str = ""):
    conn = get_db()
    rows = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC"
    ).fetchall()]
    conn.close()
    grouped = {}
    for d in rows:
        y, m, _ = d.split('-')
        grouped.setdefault(y, {}).setdefault(m, []).append(d)
    return templates.TemplateResponse(request, "dashboard/import.html", {
        "request": request, "grouped_dates": grouped, "msg": msg
    })


@dashboard_router.post("/api/import")
async def dashboard_api_import(file: UploadFile = File(...)):
    content = await file.read()
    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    import openpyxl
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    fname = file.filename or "upload.xlsx"
    wb.close()
    os.unlink(tmp_path)

    d = extract_date_from_filename(fname)
    if not d:
        d = date.today().isoformat()

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp2:
        tmp2.write(content)
        tmp2_path = tmp2.name

    cnt = import_excel(tmp2_path, d)
    os.unlink(tmp2_path)

    if cnt > 0:
        return RedirectResponse(f"/dashboard/import?msg=成功导入 {cnt} 条店铺数据（{d}）", status_code=302)
    else:
        return RedirectResponse(f"/dashboard/import?msg=该日期数据已存在或文件为空（{d}）", status_code=302)


# ==================== Task Tracker Routes ====================

@task_router.get("/", response_class=HTMLResponse)
def task_index(request: Request):
    conn = get_db()
    today = date.today().isoformat()

    shops = [dict(r) for r in conn.execute(
        "SELECT * FROM shops WHERE status != 'closed' ORDER BY name"
    ).fetchall()]

    for shop in shops:
        ensure_daily_tasks(conn, shop['id'], today)

    # Build task_stats: each task with its shops
    task_templates = [dict(r) for r in conn.execute(
        "SELECT * FROM daily_task_templates WHERE is_active=1 ORDER BY sort_order"
    ).fetchall()]

    task_stats = []
    for tmpl in task_templates:
        task_shops = []
        for shop in shops:
            row = conn.execute(
                "SELECT is_completed FROM daily_tasks WHERE shop_id=? AND task_id=? AND task_date=?",
                (shop['id'], tmpl['id'], today)
            ).fetchone()
            is_completed = row['is_completed'] if row else 0
            task_shops.append({
                'name': shop['name'],
                'group_name': shop.get('group_name', ''),
                'is_completed': is_completed,
                'shop_id': shop['id']
            })
        done_count = sum(1 for s in task_shops if s['is_completed'])
        task_stats.append({
            'task_id': tmpl['id'],
            'task_name': tmpl['task_name'],
            'total_count': len(task_shops),
            'done_count': done_count,
            'shops': task_shops
        })

    total_shops = len(shops)
    # Count shops where all tasks are done
    total_done = 0
    for shop in shops:
        total = conn.execute(
            "SELECT COUNT(*) FROM daily_tasks WHERE shop_id=? AND task_date=?",
            (shop['id'], today)
        ).fetchone()[0]
        done = conn.execute(
            "SELECT COUNT(*) FROM daily_tasks WHERE shop_id=? AND task_date=? AND is_completed=1",
            (shop['id'], today)
        ).fetchone()[0]
        if total > 0 and done == total:
            total_done += 1
        shop['total_tasks'] = total
        shop['done_tasks'] = done
        shop['progress'] = round(done / total * 100) if total > 0 else 0

        if shop['status'] == 'new':
            new_total = conn.execute(
                "SELECT COUNT(*) FROM new_shop_tasks WHERE shop_id=?", (shop['id'],)
            ).fetchone()[0]
            new_done = conn.execute(
                "SELECT COUNT(*) FROM new_shop_tasks WHERE shop_id=? AND is_completed=1",
                (shop['id'],)
            ).fetchone()[0]
            shop['new_total'] = new_total
            shop['new_done'] = new_done
            shop['new_progress'] = round(new_done / new_total * 100) if new_total > 0 else 0
        else:
            shop['new_total'] = 0
            shop['new_done'] = 0
            shop['new_progress'] = 0

    new_shops = [s for s in shops if s['status'] == 'new']

    all_groups = sorted(set(s.get('group_name', '') for s in shops if s.get('group_name')))

    conn.close()
    return templates.TemplateResponse(request, "task/index.html", {
        "request": request, "shops": shops, "today": today,
        "total_shops": total_shops, "total_done": total_done,
        "new_shops": new_shops, "task_stats": task_stats,
        "all_groups": all_groups
    })


@task_router.get("/licenses", response_class=HTMLResponse)
def task_licenses_page(request: Request):
    conn = get_db()
    licenses = [dict(r) for r in conn.execute("SELECT * FROM licenses ORDER BY created_at DESC").fetchall()]
    for lic in licenses:
        count = conn.execute("SELECT COUNT(*) FROM shops WHERE license_id=?", (lic['id'],)).fetchone()[0]
        lic['shop_count'] = count
    conn.close()
    return templates.TemplateResponse(request, "task/licenses.html", {
        "request": request, "licenses": licenses
    })


@task_router.post("/api/licenses/add")
def task_add_license(request: Request, license_name: str = Form(...), license_no: str = Form(""),
                holder_name: str = Form(""), expire_date: str = Form(""), remark: str = Form("")):
    conn = get_db()
    conn.execute(
        "INSERT INTO licenses(license_name, license_no, holder_name, expire_date, remark) VALUES(?,?,?,?,?)",
        (license_name, license_no, holder_name, expire_date, remark)
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/task/licenses", status_code=302)


@task_router.post("/api/licenses/{license_id}/delete")
def task_delete_license(license_id: int):
    conn = get_db()
    conn.execute("UPDATE shops SET license_id=NULL WHERE license_id=?", (license_id,))
    conn.execute("DELETE FROM licenses WHERE id=?", (license_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/licenses", status_code=302)


@task_router.post("/api/licenses/{license_id}/edit")
def task_edit_license(license_id: int, license_name: str = Form(...), license_no: str = Form(""),
                 holder_name: str = Form(""), expire_date: str = Form(""), remark: str = Form("")):
    conn = get_db()
    conn.execute(
        "UPDATE licenses SET license_name=?, license_no=?, holder_name=?, expire_date=?, remark=? WHERE id=?",
        (license_name, license_no, holder_name, expire_date, remark, license_id)
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/task/licenses", status_code=302)


@task_router.get("/shops", response_class=HTMLResponse)
def task_shops_page(request: Request, group: str = ""):
    conn = get_db()
    query = "SELECT s.*, l.license_name FROM shops s LEFT JOIN licenses l ON s.license_id=l.id"
    params = []
    if group:
        query += " WHERE s.group_name=?"
        params.append(group)
    query += " ORDER BY s.status, s.name"
    shops = [dict(r) for r in conn.execute(query, params).fetchall()]

    all_groups = [r['group_name'] for r in conn.execute(
        "SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name"
    ).fetchall()]
    licenses = [dict(r) for r in conn.execute("SELECT id, license_name FROM licenses ORDER BY license_name").fetchall()]
    conn.close()
    return templates.TemplateResponse(request, "task/shops.html", {
        "request": request, "shops": shops, "groups": all_groups,
        "licenses": licenses, "selected_group": group
    })


@task_router.post("/api/shops/add")
def task_add_shop(request: Request, shop_name: str = Form(...), license_id: str = Form(""),
             group_name: str = Form(""), status: str = Form("new")):
    conn = get_db()
    lid = int(license_id) if license_id else None
    try:
        cur = conn.execute(
            "INSERT INTO shops(name, license_id, group_name, status) VALUES(?,?,?,?)",
            (shop_name, lid, group_name, status)
        )
        shop_id = cur.lastrowid
        if status == 'new':
            ensure_new_shop_tasks(conn, shop_id)
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse("/task/shops", status_code=302)


@task_router.post("/api/shops/{shop_id}/edit")
def task_edit_shop(shop_id: int, shop_name: str = Form(...), license_id: str = Form(""),
              group_name: str = Form(""), status: str = Form("new")):
    conn = get_db()
    lid = int(license_id) if license_id else None
    old = conn.execute("SELECT status FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.execute(
        "UPDATE shops SET name=?, license_id=?, group_name=?, status=? WHERE id=?",
        (shop_name, lid, group_name, status, shop_id)
    )
    if old and old['status'] != 'new' and status == 'new':
        ensure_new_shop_tasks(conn, shop_id)
    if status == 'new':
        shop = conn.execute("SELECT new_shop_task_done FROM shops WHERE id=?", (shop_id,)).fetchone()
        if shop and not shop['new_shop_task_done']:
            ensure_new_shop_tasks(conn, shop_id)
    conn.commit()
    conn.close()
    return RedirectResponse("/task/shops", status_code=302)


@task_router.post("/api/shops/{shop_id}/delete")
def task_delete_shop(shop_id: int):
    conn = get_db()
    conn.execute("DELETE FROM daily_tasks WHERE shop_id=?", (shop_id,))
    conn.execute("DELETE FROM new_shop_tasks WHERE shop_id=?", (shop_id,))
    conn.execute("DELETE FROM platform_accounts WHERE shop_id=?", (shop_id,))
    conn.execute("DELETE FROM shops WHERE id=?", (shop_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/shops", status_code=302)


@task_router.get("/accounts", response_class=HTMLResponse)
def task_accounts_page(request: Request, shop_id: str = "", platform: str = ""):
    conn = get_db()
    shops = [dict(r) for r in conn.execute("SELECT id, name as shop_name FROM shops ORDER BY name").fetchall()]

    accounts = []
    if shop_id:
        rows = conn.execute(
            "SELECT a.*, s.name as shop_name FROM platform_accounts a JOIN shops s ON a.shop_id=s.id WHERE a.shop_id=? ORDER BY a.platform_name",
            (int(shop_id),)
        ).fetchall()
        accounts = [dict(r) for r in rows]
    elif platform:
        rows = conn.execute(
            "SELECT a.*, s.name as shop_name FROM platform_accounts a JOIN shops s ON a.shop_id=s.id WHERE a.platform_name=? ORDER BY s.name",
            (platform,)
        ).fetchall()
        accounts = [dict(r) for r in rows]
    else:
        rows = conn.execute(
            "SELECT a.*, s.name as shop_name FROM platform_accounts a JOIN shops s ON a.shop_id=s.id ORDER BY s.name, a.platform_name"
        ).fetchall()
        accounts = [dict(r) for r in rows]

    for acc in accounts:
        acc['password_dec'] = crypto.decrypt(acc['password_enc'])

    all_platforms = [r['platform_name'] for r in conn.execute(
        "SELECT DISTINCT platform_name FROM platform_accounts ORDER BY platform_name"
    ).fetchall()]

    conn.close()
    return templates.TemplateResponse(request, "task/accounts.html", {
        "request": request, "accounts": accounts, "shops": shops,
        "all_platforms": all_platforms, "selected_shop": shop_id, "selected_platform": platform
    })


@task_router.post("/api/accounts/add")
def task_add_account(request: Request, shop_id: int = Form(...), platform_name: str = Form(...),
                account: str = Form(""), password: str = Form(""), remark: str = Form("")):
    conn = get_db()
    enc_pwd = crypto.encrypt(password)
    conn.execute(
        "INSERT INTO platform_accounts(shop_id, platform_name, account, password_enc, remark) VALUES(?,?,?,?,?)",
        (shop_id, platform_name, account, enc_pwd, remark)
    )
    conn.commit()
    conn.close()
    return RedirectResponse(f"/task/accounts?shop_id={shop_id}", status_code=302)


@task_router.post("/api/accounts/{acc_id}/edit")
def task_edit_account(acc_id: int, platform_name: str = Form(...), account: str = Form(""),
                 password: str = Form(""), remark: str = Form("")):
    conn = get_db()
    enc_pwd = crypto.encrypt(password)
    conn.execute(
        "UPDATE platform_accounts SET platform_name=?, account=?, password_enc=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
        (platform_name, account, enc_pwd, remark, acc_id)
    )
    conn.commit()
    acc = conn.execute("SELECT shop_id FROM platform_accounts WHERE id=?", (acc_id,)).fetchone()
    conn.close()
    shop_id = acc['shop_id'] if acc else ""
    return RedirectResponse(f"/task/accounts?shop_id={shop_id}", status_code=302)


@task_router.post("/api/accounts/{acc_id}/delete")
def task_delete_account(acc_id: int):
    conn = get_db()
    acc = conn.execute("SELECT shop_id FROM platform_accounts WHERE id=?", (acc_id,)).fetchone()
    conn.execute("DELETE FROM platform_accounts WHERE id=?", (acc_id,))
    conn.commit()
    shop_id = acc['shop_id'] if acc else ""
    conn.close()
    return RedirectResponse(f"/task/accounts?shop_id={shop_id}", status_code=302)


@task_router.get("/api/accounts/{acc_id}/get_password")
def task_get_password(acc_id: int):
    conn = get_db()
    acc = conn.execute("SELECT password_enc FROM platform_accounts WHERE id=?", (acc_id,)).fetchone()
    conn.close()
    if acc:
        return {"password": crypto.decrypt(acc['password_enc'])}
    return {"password": ""}


@task_router.get("/daily_task_templates", response_class=HTMLResponse)
def task_daily_task_templates_page(request: Request):
    conn = get_db()
    tasks = [dict(r) for r in conn.execute(
        "SELECT * FROM daily_task_templates ORDER BY sort_order"
    ).fetchall()]
    conn.close()
    return templates.TemplateResponse(request, "task/daily_tasks.html", {
        "request": request, "tasks": tasks
    })


@task_router.post("/api/daily_task_templates/add")
def task_add_daily_task(task_name: str = Form(...)):
    conn = get_db()
    max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM daily_task_templates").fetchone()[0]
    conn.execute("INSERT INTO daily_task_templates(task_name, sort_order) VALUES(?,?)",
                 (task_name, max_order + 1))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/daily_task_templates", status_code=302)


@task_router.post("/api/daily_task_templates/{task_id}/toggle")
def task_toggle_daily_task(task_id: int):
    conn = get_db()
    conn.execute("UPDATE daily_task_templates SET is_active = 1 - is_active WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/daily_task_templates", status_code=302)


@task_router.post("/api/daily_task_templates/{task_id}/delete")
def task_delete_daily_task(task_id: int):
    conn = get_db()
    conn.execute("DELETE FROM daily_task_templates WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/daily_task_templates", status_code=302)


@task_router.get("/new_shop_task_templates", response_class=HTMLResponse)
def task_new_shop_task_templates_page(request: Request):
    conn = get_db()
    tasks = [dict(r) for r in conn.execute(
        "SELECT * FROM new_shop_task_templates ORDER BY sort_order"
    ).fetchall()]
    conn.close()
    return templates.TemplateResponse(request, "task/new_shop_tasks.html", {
        "request": request, "tasks": tasks
    })


@task_router.post("/api/new_shop_task_templates/add")
def task_add_new_shop_task(task_name: str = Form(...)):
    conn = get_db()
    max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM new_shop_task_templates").fetchone()[0]
    conn.execute("INSERT INTO new_shop_task_templates(task_name, sort_order) VALUES(?,?)",
                 (task_name, max_order + 1))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/new_shop_task_templates", status_code=302)


@task_router.post("/api/new_shop_task_templates/{task_id}/toggle")
def task_toggle_new_shop_task(task_id: int):
    conn = get_db()
    conn.execute("UPDATE new_shop_task_templates SET is_active = 1 - is_active WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/new_shop_task_templates", status_code=302)


@task_router.post("/api/new_shop_task_templates/{task_id}/delete")
def task_delete_new_shop_task(task_id: int):
    conn = get_db()
    conn.execute("DELETE FROM new_shop_task_templates WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/task/new_shop_task_templates", status_code=302)


@task_router.get("/checkin", response_class=HTMLResponse)
def task_checkin_page(request: Request, shop_id: str = "", qdate: str = ""):
    conn = get_db()
    today = date.today().isoformat()
    selected_date = qdate if qdate else today

    shops = [dict(r) for r in conn.execute(
        "SELECT id, name, group_name, status FROM shops WHERE status != 'closed' ORDER BY name"
    ).fetchall()]

    tasks_data = []
    selected_shop = None
    if shop_id:
        selected_shop = next((s for s in shops if s['id'] == int(shop_id)), None)
        ensure_daily_tasks(conn, int(shop_id), selected_date)
        rows = conn.execute("""
            SELECT dt.id, dt.is_completed, dt.completed_at, dt.remark,
                   t.task_name, t.sort_order
            FROM daily_tasks dt
            JOIN daily_task_templates t ON dt.task_id = t.id
            WHERE dt.shop_id=? AND dt.task_date=?
            ORDER BY t.sort_order
        """, (int(shop_id), selected_date)).fetchall()
        tasks_data = [dict(r) for r in rows]

    conn.close()
    return templates.TemplateResponse(request, "task/checkin.html", {
        "request": request, "shops": shops, "tasks": tasks_data,
        "selected_shop": selected_shop, "selected_date": selected_date, "today": today
    })


@task_router.post("/api/checkin/toggle")
def task_checkin_toggle(shop_id: int = Form(...), task_id: int = Form(...), task_date: str = Form(...)):
    conn = get_db()
    row = conn.execute(
        "SELECT is_completed FROM daily_tasks WHERE shop_id=? AND task_id=? AND task_date=?",
        (shop_id, task_id, task_date)
    ).fetchone()
    if row:
        new_val = 0 if row['is_completed'] else 1
        completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if new_val else ""
        conn.execute(
            "UPDATE daily_tasks SET is_completed=?, completed_at=? WHERE shop_id=? AND task_id=? AND task_date=?",
            (new_val, completed_at, shop_id, task_id, task_date)
        )
    conn.commit()
    conn.close()
    return {"ok": True}


@task_router.post("/api/checkin/complete_all")
def task_checkin_complete_all(shop_id: int = Form(...), task_date: str = Form(...)):
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        "UPDATE daily_tasks SET is_completed=1, completed_at=? WHERE shop_id=? AND task_date=?",
        (now, shop_id, task_date)
    )
    conn.commit()
    conn.close()
    return {"ok": True}


@task_router.get("/new_shop_checkin", response_class=HTMLResponse)
def task_new_shop_checkin_page(request: Request, shop_id: str = ""):
    conn = get_db()
    new_shops = [dict(r) for r in conn.execute(
        "SELECT id, name as shop_name FROM shops WHERE status='new' ORDER BY name"
    ).fetchall()]

    tasks_data = []
    selected_shop = None
    if shop_id:
        selected_shop = next((s for s in new_shops if s['id'] == int(shop_id)), None)
        ensure_new_shop_tasks(conn, int(shop_id))
        rows = conn.execute("""
            SELECT ns.id, ns.is_completed, ns.completed_at, ns.remark,
                   t.task_name, t.sort_order
            FROM new_shop_tasks ns
            JOIN new_shop_task_templates t ON ns.task_id = t.id
            WHERE ns.shop_id=?
            ORDER BY t.sort_order
        """, (int(shop_id),)).fetchall()
        tasks_data = [dict(r) for r in rows]

    conn.close()
    return templates.TemplateResponse(request, "task/new_shop_checkin.html", {
        "request": request, "new_shops": new_shops, "tasks": tasks_data,
        "selected_shop": selected_shop
    })


@task_router.post("/api/new_shop_checkin/toggle")
def task_new_shop_checkin_toggle(shop_id: int = Form(...), task_id: int = Form(...)):
    conn = get_db()
    row = conn.execute(
        "SELECT is_completed FROM new_shop_tasks WHERE shop_id=? AND task_id=?",
        (shop_id, task_id)
    ).fetchone()
    if row:
        new_val = 0 if row['is_completed'] else 1
        completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if new_val else ""
        conn.execute(
            "UPDATE new_shop_tasks SET is_completed=?, completed_at=? WHERE shop_id=? AND task_id=?",
            (new_val, completed_at, shop_id, task_id)
        )
    conn.commit()
    conn.close()
    return {"ok": True}


@task_router.post("/api/new_shop_checkin/complete_all")
def task_new_shop_checkin_complete_all(shop_id: int = Form(...)):
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        "UPDATE new_shop_tasks SET is_completed=1, completed_at=? WHERE shop_id=?",
        (now, shop_id)
    )
    conn.execute("UPDATE shops SET new_shop_task_done=1, status='normal' WHERE id=?", (shop_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@task_router.get("/new_shop_history", response_class=HTMLResponse)
def task_new_shop_history_page(request: Request, shop_id: str = ""):
    conn = get_db()
    completed_shops = [dict(r) for r in conn.execute("""
        SELECT id, name as shop_name, created_at
        FROM shops
        WHERE new_shop_task_done=1
        ORDER BY name
    """).fetchall()]

    tasks_data = []
    selected_shop = None
    if shop_id:
        selected_shop = next((s for s in completed_shops if s['id'] == int(shop_id)), None)
        rows = conn.execute("""
            SELECT ns.is_completed, ns.completed_at, ns.remark,
                   t.task_name, t.sort_order
            FROM new_shop_tasks ns
            JOIN new_shop_task_templates t ON ns.task_id = t.id
            WHERE ns.shop_id=?
            ORDER BY t.sort_order
        """, (int(shop_id),)).fetchall()
        tasks_data = [dict(r) for r in rows]

    conn.close()
    return templates.TemplateResponse(request, "task/new_shop_history.html", {
        "request": request, "completed_shops": completed_shops,
        "tasks": tasks_data, "selected_shop": selected_shop
    })


@task_router.get("/stats", response_class=HTMLResponse)
def task_stats_page(request: Request, days: int = 7):
    conn = get_db()
    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    daily_stats = [dict(r) for r in conn.execute("""
        SELECT task_date,
               COUNT(DISTINCT shop_id) as total_shops,
               SUM(CASE WHEN is_completed=1 THEN 1 ELSE 0 END) as done_count,
               COUNT(*) as total_count
        FROM daily_tasks
        WHERE task_date >= ? AND task_date <= ?
        GROUP BY task_date ORDER BY task_date
    """, (start_date, today)).fetchall()]

    for d in daily_stats:
        d['completion_rate'] = round(d['done_count'] / d['total_count'] * 100) if d['total_count'] > 0 else 0

    new_shops_progress = [dict(r) for r in conn.execute("""
        SELECT s.name as shop_name,
               COUNT(*) as total,
               SUM(CASE WHEN ns.is_completed=1 THEN 1 ELSE 0 END) as done
        FROM new_shop_tasks ns
        JOIN shops s ON ns.shop_id=s.id
        WHERE s.status='new'
        GROUP BY s.id
    """).fetchall()]

    for ns in new_shops_progress:
        ns['progress'] = round(ns['done'] / ns['total'] * 100) if ns['total'] > 0 else 0

    all_groups = [r['group_name'] for r in conn.execute(
        "SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name"
    ).fetchall()]

    # Group statistics
    group_stats = [dict(r) for r in conn.execute("""
        SELECT s.group_name,
               COUNT(DISTINCT s.id) as shop_count,
               COUNT(dt.id) as total_tasks,
               SUM(CASE WHEN dt.is_completed=1 THEN 1 ELSE 0 END) as done_tasks
        FROM shops s
        LEFT JOIN daily_tasks dt ON s.id = dt.shop_id AND dt.task_date >= ? AND dt.task_date <= ?
        WHERE s.group_name != '' AND s.status != 'closed'
        GROUP BY s.group_name
        ORDER BY s.group_name
    """, (start_date, today)).fetchall()]

    for g in group_stats:
        g['completion_rate'] = round(g['done_tasks'] / g['total_tasks'] * 100) if g['total_tasks'] > 0 else 0

    license_stats = [dict(r) for r in conn.execute("""
        SELECT l.license_name, COUNT(s.id) as shop_count
        FROM licenses l LEFT JOIN shops s ON s.license_id=l.id
        GROUP BY l.id ORDER BY shop_count DESC
    """).fetchall()]

    conn.close()
    return templates.TemplateResponse(request, "task/stats.html", {
        "request": request, "daily_stats": daily_stats, "days": days,
        "new_shops_progress": new_shops_progress, "license_stats": license_stats,
        "all_groups": all_groups, "today": today, "group_stats": group_stats
    })


# ==================== Main App Routes ====================

@app.get("/", response_class=HTMLResponse)
def main_index(request: Request):
    return RedirectResponse("/profit/", status_code=302)


@app.get("/backup", response_class=HTMLResponse)
def backup_page(request: Request):
    return templates.TemplateResponse(request, "backup.html", {"request": request})


@app.get("/api/backup/download")
def backup_download():
    import zipfile
    import tempfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        if os.path.exists(DB_PATH):
            zf.write(DB_PATH, "data.db")
        if os.path.exists(os.path.join(BASE_DIR, "secret.key")):
            zf.write(os.path.join(BASE_DIR, "secret.key"), "secret.key")
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=shop_backup.zip"})


@app.post("/api/backup/restore")
async def backup_restore(file: UploadFile = File(...)):
    import zipfile
    content = await file.read()
    with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
        names = zf.namelist()
        if "data.db" in names:
            zf.extract("data.db", BASE_DIR)
        if "secret.key" in names:
            zf.extract("secret.key", BASE_DIR)
    return RedirectResponse("/backup?msg=数据恢复成功，请重启应用", status_code=302)


# ==================== Register Routers ====================

app.include_router(profit_router)
app.include_router(dashboard_router)
app.include_router(task_router)


def create_tray_icon():
    import pystray
    from PIL import Image, ImageDraw

    img = Image.new('RGBA', (64, 64), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.ellipse([4, 4, 60, 60], fill=(0, 122, 255))
    draw.text((16, 12), "SM", fill=(255, 255, 255))

    def on_open(icon, item):
        import webbrowser
        webbrowser.open("http://localhost:8000")

    def on_exit(icon, item):
        icon.stop()
        import os
        os._exit(0)

    menu = pystray.Menu(
        pystray.MenuItem("Open ShopManager", on_open, default=True),
        pystray.MenuItem("Exit", on_exit)
    )
    icon = pystray.Icon("ShopManager", img, "ShopManager", menu)
    return icon


if __name__ == "__main__":
    import uvicorn
    import threading

    tray_icon = create_tray_icon()
    tray_thread = threading.Thread(target=tray_icon.run, daemon=True)
    tray_thread.start()

    uvicorn.run(app, host="0.0.0.0", port=8000, timeout_keep_alive=120)
