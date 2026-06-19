import os
import io
import sqlite3
from datetime import date, timedelta
from typing import Optional

from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data.db")

app = FastAPI()
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def calc_profit(douyin_amount, taobao_amount, refund_status, is_influencer, warehouse_status):
    """计算利润"""
    profit = douyin_amount - taobao_amount
    # 运费6.9：未到仓库+退款或已到仓未发货+退款时退给用户，其他情况扣除
    if refund_status in ('已发货退款', '退货退款'):
        profit -= 6.9  # 运费不退，扣除
    elif refund_status not in ('待发货',) and warehouse_status == '未到仓库':
        pass  # 未到仓库+退款，运费退给用户，不扣除
    elif refund_status not in ('待发货',) and warehouse_status == '已到达仓库未发货':
        pass  # 已到仓未发货+退款，运费退给用户，不扣除
    else:
        profit -= 6.9  # 正常订单扣除运费
    # 达人佣金：任何退款都不扣
    if is_influencer == '是' and refund_status in ('待发货', '已发货'):
        profit -= douyin_amount * 0.25
    # 发货成本5.5：仅已到仓的订单
    if warehouse_status in ('已到达仓库未发货', '已到仓库已发货'):
        if refund_status in ('待发货', '已发货'):
            profit -= 5.5  # 正常订单扣除发货成本
    # 退货成本3.5：已到仓未发货+任何退款 或 已发货退款/退货退款
    if warehouse_status == '已到达仓库未发货' and refund_status not in ('待发货', '已发货'):
        profit -= 3.5
    elif refund_status in ('已发货退款', '退货退款'):
        profit -= 3.5
    return profit


@app.on_event("startup")
def startup():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS shops (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, group_name TEXT DEFAULT '');
        CREATE TABLE IF NOT EXISTS orders (
            douyin_order_no TEXT PRIMARY KEY,
            shop_id INTEGER NOT NULL,
            product_name TEXT NOT NULL,
            douyin_amount REAL NOT NULL,
            taobao_order_no TEXT DEFAULT '',
            taobao_amount REAL DEFAULT 0,
            refund_status TEXT DEFAULT '待发货',
            order_date TEXT NOT NULL,
            buyer_note TEXT DEFAULT '',
            system_note TEXT DEFAULT '',
            logistics_company TEXT DEFAULT '',
            logistics_no TEXT DEFAULT '',
            is_influencer TEXT DEFAULT '否',
            warehouse_status TEXT DEFAULT '未到仓库'
        );
    """)
    # Add new columns if missing
    for col, typ in [("buyer_note", "TEXT DEFAULT ''"), ("system_note", "TEXT DEFAULT ''"),
                     ("logistics_company", "TEXT DEFAULT ''"), ("logistics_no", "TEXT DEFAULT ''"),
                     ("is_influencer", "TEXT DEFAULT '否'"), ("warehouse_status", "TEXT DEFAULT '未到仓库'")]:
        try: conn.execute(f"ALTER TABLE orders ADD COLUMN {col} {typ}")
        except: pass
    try: conn.execute("ALTER TABLE shops ADD COLUMN group_name TEXT DEFAULT ''")
    except: pass
    if not conn.execute("SELECT id FROM shops WHERE name='默认店铺'").fetchone():
        conn.execute("INSERT INTO shops(name) VALUES('默认店铺')")
    conn.commit()
    conn.close()


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    db = get_db()
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    month_start = date.today().replace(day=1).isoformat()

    def sf(d):
        return dict(db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(douyin_amount),0) as dy, "
            "COALESCE(SUM(taobao_amount),0) as tb, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
            "ELSE 6.9 END),0) as freight, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END),0) as commission, "
            "COALESCE(SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END),0) as ship_cost, "
            "COALESCE(SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
            "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END),0) as return_cost "
            "FROM orders WHERE order_date=?", (d,)
        ).fetchone())

    def sf2(a, b):
        return dict(db.execute(
            "SELECT COUNT(*) as cnt, COALESCE(SUM(douyin_amount),0) as dy, "
            "COALESCE(SUM(taobao_amount),0) as tb, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
            "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
            "ELSE 6.9 END),0) as freight, "
            "COALESCE(SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END),0) as commission, "
            "COALESCE(SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END),0) as ship_cost, "
            "COALESCE(SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
            "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END),0) as return_cost "
            "FROM orders WHERE order_date>=? AND order_date<=?", (a, b)
        ).fetchone())

    td = sf(today); td["profit"] = td["dy"] - td["tb"] - td["freight"] - td["commission"] - td["ship_cost"] - td["return_cost"]
    wk = sf2(week_ago, today); wk["profit"] = wk["dy"] - wk["tb"] - wk["freight"] - wk["commission"] - wk["ship_cost"] - wk["return_cost"]
    mn = sf2(month_start, today); mn["profit"] = mn["dy"] - mn["tb"] - mn["freight"] - mn["commission"] - mn["ship_cost"] - mn["return_cost"]

    trend = [dict(r) for r in db.execute(
        "SELECT order_date, "
        "SUM(douyin_amount)-SUM(taobao_amount)-"
        "SUM(CASE WHEN refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        "WHEN refund_status NOT IN ('待发货') AND warehouse_status='未到仓库' THEN 0 "
        "WHEN refund_status NOT IN ('待发货') AND warehouse_status='已到达仓库未发货' THEN 0 "
        "ELSE 6.9 END)-"
        "SUM(CASE WHEN refund_status IN ('待发货','已发货') AND is_influencer='是' THEN douyin_amount*0.25 ELSE 0 END)-"
        "SUM(CASE WHEN warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        "SUM(CASE WHEN warehouse_status='已到达仓库未发货' AND refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        "WHEN refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        "FROM orders WHERE order_date>=date('now','-7 days') GROUP BY order_date ORDER BY order_date"
    ).fetchall()]

    sp = [dict(r) for r in db.execute(
        "SELECT s.name, s.id, s.group_name, "
        "SUM(o.douyin_amount)-SUM(o.taobao_amount)-"
        "SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        "WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        "WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        "ELSE 6.9 END)-"
        "SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END)-"
        "SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        "SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        "WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        "FROM orders o JOIN shops s ON o.shop_id=s.id WHERE o.order_date>=date('now','-30 days') "
        "GROUP BY o.shop_id ORDER BY profit DESC"
    ).fetchall()]
    
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    
    db.close()
    return templates.TemplateResponse(request, "index.html", {"td": td, "wk": wk, "mn": mn, "trend": trend, "shop_profit": sp, "groups": [g["group_name"] for g in groups]})


@app.get("/api/trend")
def api_trend(date_from: Optional[str] = None, date_to: Optional[str] = None, shop_id: Optional[int] = None, group: Optional[str] = None):
    db = get_db()
    df = date_from or date.today().isoformat()
    dt = date_to or date.today().isoformat()
    c, p = ["o.order_date>=?", "o.order_date<=?"], [df, dt]
    if shop_id: c.append("o.shop_id=?"); p.append(shop_id)
    if group: c.append("s.group_name=?"); p.append(group)
    w = " AND ".join(c)
    trend = [dict(r) for r in db.execute(
        f"SELECT o.order_date, "
        f"SUM(o.douyin_amount)-SUM(o.taobao_amount)-"
        f"SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        f"ELSE 6.9 END)-"
        f"SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END)-"
        f"SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END)-"
        f"SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        f"WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as profit "
        f"FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE {w} GROUP BY o.order_date ORDER BY o.order_date", p
    ).fetchall()]
    db.close()
    return trend


@app.get("/orders", response_class=HTMLResponse)
def orders_page(request: Request, shop_id: str = "", date_from: str = "",
                date_to: str = "", status: str = "",
                keyword: str = "", page: int = 1, group: str = ""):
    db = get_db()
    pp = 20
    c, p = [], []
    if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
    if date_from: c.append("o.order_date>=?"); p.append(date_from)
    if date_to: c.append("o.order_date<=?"); p.append(date_to)
    if status: c.append("o.refund_status=?"); p.append(status)
    if keyword: c.append("(o.douyin_order_no LIKE ? OR o.product_name LIKE ? OR o.taobao_order_no LIKE ?)"); lk = f"%{keyword}%"; p.extend([lk,lk,lk])
    if group: c.append("s.group_name=?"); p.append(group)
    w = ("WHERE " + " AND ".join(c)) if c else ""
    total = db.execute(f"SELECT COUNT(*) FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w}", p).fetchone()[0]
    items = [dict(r) for r in db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w} ORDER BY o.order_date DESC, o.douyin_order_no DESC LIMIT ? OFFSET ?", p + [pp, (page-1)*pp]).fetchall()]
    shops = [dict(r) for r in db.execute("SELECT * FROM shops ORDER BY id").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    today = date.today().isoformat()
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    d7 = (date.today() - timedelta(days=6)).isoformat()
    d15 = (date.today() - timedelta(days=14)).isoformat()
    d30 = (date.today() - timedelta(days=29)).isoformat()
    return templates.TemplateResponse(request, "orders.html", {"items": items, "shops": shops, "total": total, "page": page, "per_page": pp, "f_sid": shop_id, "f_df": date_from, "f_dt": date_to, "f_st": status, "f_kw": keyword, "f_group": group, "today": today, "yesterday": yesterday, "d7": d7, "d15": d15, "d30": d30, "groups": [g["group_name"] for g in groups]})


@app.get("/orders/new", response_class=HTMLResponse)
def order_new(request: Request):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "order_form.html", {"order": None, "shops": shops})


@app.get("/orders/export")
def export_orders(shop_id: str = "", date_from: str = "",
                  date_to: str = "", status: str = "",
                  keyword: str = "", selected: str = "", group: str = ""):
    db = get_db()
    if selected:
        order_list = selected.split(",")
        placeholders = ",".join(["?" for _ in order_list])
        rows = db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no IN ({placeholders}) ORDER BY o.order_date DESC", order_list).fetchall()
    else:
        c, p = [], []
        if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
        if date_from: c.append("o.order_date>=?"); p.append(date_from)
        if date_to: c.append("o.order_date<=?"); p.append(date_to)
        if status: c.append("o.refund_status=?"); p.append(status)
        if keyword: c.append("(o.douyin_order_no LIKE ? OR o.product_name LIKE ? OR o.taobao_order_no LIKE ?)"); lk = f"%{keyword}%"; p.extend([lk,lk,lk])
        if group: c.append("s.group_name=?"); p.append(group)
        w = ("WHERE " + " AND ".join(c)) if c else ""
        rows = db.execute(f"SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id {w} ORDER BY o.order_date DESC", p).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单数据"
    ws.append(["店铺", "抖音订单号", "商品名称", "抖音金额", "淘宝订单号", "淘宝金额", "利润", "订单状态", "订单日期", "买家备注", "系统备注", "物流公司", "物流单号", "达人带货", "仓库状态", "运费", "达人佣金", "发货成本", "退货成本"])
    for r in rows:
        freight = 6.9
        commission = r["douyin_amount"] * 0.25 if r["is_influencer"] == "是" and r["refund_status"] in ("待发货", "已发货") else 0
        ship_cost = 5.5 if r["warehouse_status"] in ("已到达仓库未发货", "已到仓库已发货") and r["refund_status"] in ("待发货", "已发货") else 0
        return_cost = 3.5 if (r["warehouse_status"] == "已到达仓库未发货" and r["refund_status"] not in ("待发货", "已发货")) or r["refund_status"] in ("已发货退款", "退货退款") else 0
        # 运费逻辑：未到仓库+退款或已到仓未发货+退款时退给用户
        if r["refund_status"] in ("已发货退款", "退货退款"):
            freight = 6.9
        elif r["refund_status"] not in ("待发货",) and r["warehouse_status"] == "未到仓库":
            freight = 0
        elif r["refund_status"] not in ("待发货",) and r["warehouse_status"] == "已到达仓库未发货":
            freight = 0
        else:
            freight = 6.9
        profit = r["douyin_amount"] - r["taobao_amount"] - freight - commission - ship_cost - return_cost
        ws.append([r["shop_name"], r["douyin_order_no"], r["product_name"], r["douyin_amount"],
                   r["taobao_order_no"], r["taobao_amount"], profit, r["refund_status"], r["order_date"],
                   r["buyer_note"], r["system_note"], r["logistics_company"], r["logistics_no"],
                   r["is_influencer"], r["warehouse_status"], freight, commission, ship_cost, return_cost])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=orders_export.xlsx"})


@app.get("/orders/{dy_no}", response_class=HTMLResponse)
def order_detail(request: Request, dy_no: str):
    db = get_db()
    order = db.execute("SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no=?", (dy_no,)).fetchone()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    if not order:
        return RedirectResponse("/orders", status_code=302)
    return templates.TemplateResponse(request, "order_detail.html", {"order": dict(order), "shops": shops})


@app.get("/api/orders/{dy_no}/json")
def api_order_json(dy_no: str):
    db = get_db()
    order = db.execute("SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.douyin_order_no=?", (dy_no,)).fetchone()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    if not order:
        return {"error": "not found"}
    return {"order": dict(order), "shops": shops}


@app.get("/api/orders/by-date/{date_str}")
def api_orders_by_date(date_str: str):
    db = get_db()
    orders = [dict(r) for r in db.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.order_date=? ORDER BY o.douyin_order_no", (date_str,)
    ).fetchall()]
    for o in orders:
        o["profit"] = calc_profit(o["douyin_amount"], o["taobao_amount"], o["refund_status"], o["is_influencer"], o["warehouse_status"])
    db.close()
    return {"orders": orders}


@app.get("/api/shop/{sid}/orders")
def api_shop_orders(sid: int):
    db = get_db()
    shop = db.execute("SELECT * FROM shops WHERE id=?", (sid,)).fetchone()
    orders = [dict(r) for r in db.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.shop_id=? ORDER BY o.order_date DESC", (sid,)
    ).fetchall()]
    for o in orders:
        o["profit"] = calc_profit(o["douyin_amount"], o["taobao_amount"], o["refund_status"], o["is_influencer"], o["warehouse_status"])
    db.close()
    return {"shop": dict(shop) if shop else None, "orders": orders}


@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request, msg: Optional[str] = None):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "import.html", {"shops": shops, "msg": msg or ""})


@app.get("/shops", response_class=HTMLResponse)
def shops_page(request: Request, msg: Optional[str] = None):
    db = get_db()
    shops = [dict(r) for r in db.execute("SELECT * FROM shops ORDER BY id").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "shops.html", {"shops": shops, "msg": msg or "", "groups": [g["group_name"] for g in groups]})


@app.post("/api/shops/add")
def api_add_shop(name: str = Form(...), group_name: str = Form("")):
    db = get_db()
    try:
        db.execute("INSERT INTO shops(name, group_name) VALUES(?, ?)", (name, group_name))
        db.commit()
    except sqlite3.IntegrityError:
        pass
    db.close()
    return RedirectResponse("/shops?msg=店铺添加成功", status_code=302)


@app.post("/api/shops/{sid}/update")
def api_update_shop(sid: int, name: str = Form(...), group_name: str = Form("")):
    db = get_db()
    db.execute("UPDATE shops SET name=?, group_name=? WHERE id=?", (name, group_name, sid))
    db.commit()
    db.close()
    return RedirectResponse("/shops", status_code=302)


@app.get("/api/groups")
def api_get_groups():
    db = get_db()
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return [g["group_name"] for g in groups]


@app.post("/api/shops/{sid}/delete")
def api_delete_shop(sid: int):
    db = get_db()
    db.execute("DELETE FROM shops WHERE id=?", (sid,))
    db.commit()
    db.close()
    return RedirectResponse("/shops?msg=店铺已删除", status_code=302)


@app.get("/stats", response_class=HTMLResponse)
def stats_page(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None, shop_id: str = "", days: Optional[int] = None, group: Optional[str] = None):
    if days:
        df = (date.today() - timedelta(days=days-1)).isoformat()
    else:
        df = date_from or (date.today() - timedelta(days=29)).isoformat()
    dt = date_to or date.today().isoformat()
    db = get_db()
    c, p = ["o.order_date>=?", "o.order_date<=?"], [df, dt]
    if shop_id: c.append("o.shop_id=?"); p.append(int(shop_id))
    if group: c.append("s.group_name=?"); p.append(group)
    w = " AND ".join(c)
    daily = [dict(r) for r in db.execute(
        f"SELECT o.order_date, COUNT(*) as order_count, SUM(o.douyin_amount) as total_douyin, SUM(o.taobao_amount) as total_taobao, "
        f"SUM(CASE WHEN o.refund_status IN ('已发货退款','退货退款') THEN 6.9 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='未到仓库' THEN 0 "
        f"WHEN o.refund_status NOT IN ('待发货') AND o.warehouse_status='已到达仓库未发货' THEN 0 "
        f"ELSE 6.9 END) as total_freight, "
        f"SUM(CASE WHEN o.refund_status IN ('待发货','已发货') AND o.is_influencer='是' THEN o.douyin_amount*0.25 ELSE 0 END) as total_commission, "
        f"SUM(CASE WHEN o.warehouse_status IN ('已到达仓库未发货','已到仓库已发货') AND o.refund_status IN ('待发货','已发货') THEN 5.5 ELSE 0 END) as total_ship, "
        f"SUM(CASE WHEN o.warehouse_status='已到达仓库未发货' AND o.refund_status NOT IN ('待发货','已发货') THEN 3.5 "
        f"WHEN o.refund_status IN ('已发货退款','退货退款') THEN 3.5 ELSE 0 END) as total_return "
        f"FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE {w} GROUP BY o.order_date ORDER BY o.order_date", p
    ).fetchall()]
    for row in daily: row["profit"] = row["total_douyin"] - row["total_taobao"] - row["total_freight"] - row["total_commission"] - row["total_ship"] - row["total_return"]
    shops = [dict(r) for r in db.execute("SELECT * FROM shops").fetchall()]
    groups = [dict(r) for r in db.execute("SELECT DISTINCT group_name FROM shops WHERE group_name != '' ORDER BY group_name").fetchall()]
    db.close()
    return templates.TemplateResponse(request, "stats.html", {"daily": daily, "shops": shops, "f_df": df, "f_dt": dt, "f_sid": shop_id, "f_days": days or "", "f_group": group or "", "groups": [g["group_name"] for g in groups]})


@app.post("/api/orders/import")
async def api_import(file: UploadFile = File(...), shop_id: int = Form(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    cm = {}
    for i, h in enumerate(headers):
        if h == "订单号": cm["dy_no"] = i
        elif h == "店铺名称": cm["shop"] = i
        elif h == "下单时间": cm["date"] = i
        elif h == "订单状态": cm["status"] = i
        elif h == "买家备注": cm["buyer_note"] = i
        elif h == "系统备注": cm["sys_note"] = i
        elif h == "订单金额": cm["dy_amt"] = i
        elif h == "物流公司名称": cm["logistics"] = i
        elif h == "物流单号": cm["logistics_no"] = i
        elif h == "商品名称": cm["name"] = i
        elif h == "子订单状态": cm["sub_status"] = i
    db = get_db()
    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        def g(key, default=None):
            i = cm.get(key)
            return row[i] if i is not None and i < len(row) and row[i] is not None else default
        dy_no = str(g("dy_no", ""))
        name = str(g("name", ""))
        if not dy_no and not name: continue
        dy_amt = float(g("dy_amt", 0) or 0)
        order_date = str(g("date", date.today().isoformat()))[:10]
        status = str(g("status", ""))
        sub_status = str(g("sub_status", ""))
        if "退" in str(status) or "退" in str(sub_status):
            rf = "退货退款"
        elif "发" in str(status) or "发" in str(sub_status):
            rf = "已发货"
        else:
            rf = "待发货"
        real_shop_id = shop_id
        shop_name = g("shop")
        if shop_name:
            shop_name = str(shop_name).strip()
            row_s = db.execute("SELECT id FROM shops WHERE name=?", (shop_name,)).fetchone()
            if row_s:
                real_shop_id = row_s["id"]
            else:
                cur = db.execute("INSERT INTO shops(name, group_name) VALUES(?, '')", (shop_name,))
                real_shop_id = cur.lastrowid
        try:
            db.execute("INSERT INTO orders(douyin_order_no,shop_id,product_name,douyin_amount,taobao_order_no,taobao_amount,refund_status,order_date,buyer_note,system_note,logistics_company,logistics_no,is_influencer,warehouse_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                       (dy_no, real_shop_id, name, dy_amt, "", 0, rf, order_date,
                        str(g("buyer_note", "")), str(g("sys_note", "")), str(g("logistics", "")), str(g("logistics_no", "")), "否", "未到仓库"))
            cnt += 1
        except sqlite3.IntegrityError:
            pass
    db.commit(); db.close(); wb.close()
    return RedirectResponse(f"/import?msg=成功导入{cnt}条订单", status_code=302)


@app.post("/api/orders")
def api_create(shop_id: int = Form(...), douyin_order_no: str = Form(...), product_name: str = Form(...),
               douyin_amount: float = Form(...), taobao_order_no: str = Form(""), taobao_amount: float = Form(0),
               refund_status: str = Form("待发货"), order_date: str = Form(...),
               buyer_note: str = Form(""), system_note: str = Form(""), logistics_company: str = Form(""), logistics_no: str = Form(""),
               is_influencer: str = Form("否"), warehouse_status: str = Form("未到仓库")):
    db = get_db()
    try:
        db.execute("INSERT INTO orders(douyin_order_no,shop_id,product_name,douyin_amount,taobao_order_no,taobao_amount,refund_status,order_date,buyer_note,system_note,logistics_company,logistics_no,is_influencer,warehouse_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                   (douyin_order_no, shop_id, product_name, douyin_amount, taobao_order_no, taobao_amount, refund_status, order_date, buyer_note, system_note, logistics_company, logistics_no, is_influencer, warehouse_status))
        db.commit()
    except sqlite3.IntegrityError:
        pass
    db.close()
    return RedirectResponse("/orders", status_code=302)


@app.post("/api/orders/{dy_no}")
def api_update(dy_no: str, shop_id: int = Form(...), product_name: str = Form(...),
               douyin_amount: float = Form(...), taobao_order_no: str = Form(""), taobao_amount: float = Form(0),
               refund_status: str = Form("待发货"), order_date: str = Form(...),
               buyer_note: str = Form(""), system_note: str = Form(""), logistics_company: str = Form(""), logistics_no: str = Form(""),
               is_influencer: str = Form("否"), warehouse_status: str = Form("未到仓库")):
    db = get_db()
    db.execute("UPDATE orders SET shop_id=?,product_name=?,douyin_amount=?,taobao_order_no=?,taobao_amount=?,refund_status=?,order_date=?,buyer_note=?,system_note=?,logistics_company=?,logistics_no=?,is_influencer=?,warehouse_status=? WHERE douyin_order_no=?",
               (shop_id, product_name, douyin_amount, taobao_order_no, taobao_amount, refund_status, order_date, buyer_note, system_note, logistics_company, logistics_no, is_influencer, warehouse_status, dy_no))
    db.commit(); db.close()
    return RedirectResponse("/orders", status_code=302)


@app.post("/api/orders/{dy_no}/delete")
def api_delete(dy_no: str):
    db = get_db()
    db.execute("DELETE FROM orders WHERE douyin_order_no=?", (dy_no,))
    db.commit(); db.close()
    return RedirectResponse("/orders", status_code=302)


@app.post("/api/shops")
def api_create_shop(name: str = Form(...)):
    db = get_db()
    try: db.execute("INSERT INTO shops(name) VALUES(?)", (name,)); db.commit()
    except: pass
    db.close()
    return RedirectResponse("/", status_code=302)


@app.get("/api/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单号", "店铺名称", "下单时间", "订单状态", "买家备注", "系统备注", "订单金额", "物流公司名称", "物流单号", "商品名称", "子订单状态"])
    ws.append(["DY20240101001", "默认店铺", "2024-01-01 12:00:00", "已付款", "", "", 99.99, "顺丰速运", "SF1234567890", "示例商品", "已付款"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template.xlsx"})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, timeout_keep_alive=120)
