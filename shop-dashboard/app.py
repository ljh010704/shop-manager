import os
import re
import sqlite3
import threading
import time
from datetime import date, datetime, timedelta

import openpyxl
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_ROOT = os.path.dirname(BASE_DIR)
DB_PATH = os.path.join(BASE_DIR, "data.db")

app = FastAPI()
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


@app.on_event("startup")
def startup():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS shops (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_name TEXT NOT NULL UNIQUE,
            shop_id TEXT DEFAULT '',
            group_name TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS daily_snapshot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            status TEXT DEFAULT '',
            expire_info TEXT DEFAULT '',
            warning INTEGER DEFAULT 0,
            amount REAL DEFAULT 0,
            order_count INTEGER DEFAULT 0,
            exposure INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            refund_amount REAL DEFAULT 0,
            experience_score TEXT DEFAULT '',
            pending_ship INTEGER DEFAULT 0,
            pending_after_sale INTEGER DEFAULT 0,
            on_sale INTEGER DEFAULT 0,
            in_warehouse INTEGER DEFAULT 0,
            draft INTEGER DEFAULT 0,
            ship_24h TEXT DEFAULT '',
            overdue_ship INTEGER DEFAULT 0,
            after_sale_24h TEXT DEFAULT '',
            pending_rectify INTEGER DEFAULT 0,
            last_violation TEXT DEFAULT '',
            work_order INTEGER DEFAULT 0,
            fund REAL DEFAULT 0,
            experience_fund REAL DEFAULT 0,
            insurance_price TEXT DEFAULT '',
            valid_orders_30d INTEGER DEFAULT 0,
            cleaned_orders INTEGER DEFAULT 0,
            violation_count INTEGER DEFAULT 0,
            product_score TEXT DEFAULT '',
            logistics_score TEXT DEFAULT '',
            service_score TEXT DEFAULT '',
            abnormal_parcel INTEGER DEFAULT 0,
            recycle_bin INTEGER DEFAULT 0,
            refund_today REAL DEFAULT 0,
            qianchuan_id TEXT DEFAULT '',
            need_invoice REAL DEFAULT 0,
            today_orders INTEGER DEFAULT 0,
            remark TEXT DEFAULT '',
            UNIQUE(shop_id, snapshot_date)
        );
    """)
    conn.close()
    auto_import_all()


def safe_float(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    try:
        if isinstance(v, str):
            v = v.replace('元', '').replace('/单起', '').replace('/单', '').strip()
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_int(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def extract_date_from_filename(filename):
    m = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', filename)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def scan_excel_files():
    files = []
    if not os.path.exists(REPORT_ROOT):
        return files
    for entry in os.listdir(REPORT_ROOT):
        subdir = os.path.join(REPORT_ROOT, entry)
        if os.path.isdir(subdir) and re.match(r'\d{4}-\d{2}', entry):
            for fname in os.listdir(subdir):
                if fname.endswith('.xlsx') and not fname.startswith('~'):
                    fpath = os.path.join(subdir, fname)
                    d = extract_date_from_filename(fname)
                    if d:
                        files.append((fpath, d))
    return files


def import_excel(filepath, snapshot_date):
    conn = get_db()
    existing = conn.execute("SELECT COUNT(*) FROM daily_snapshot WHERE snapshot_date=?", (snapshot_date,)).fetchone()[0]
    if existing > 0:
        conn.close()
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value or '').strip())

    col_map = {}
    for i, h in enumerate(headers):
        if h == '店铺名称': col_map['shop_name'] = i
        elif h == '状态': col_map['status'] = i
        elif h == '到期时间': col_map['expire_info'] = i
        elif h == '预警': col_map['warning'] = i
        elif h == '成交金额': col_map['amount'] = i
        elif h == '订单数': col_map['order_count'] = i
        elif h == '曝光人数': col_map['exposure'] = i
        elif h == '点击人数': col_map['clicks'] = i
        elif h == '退款金额': col_map['refund_amount'] = i
        elif h == '体验分': col_map['experience_score'] = i
        elif h == '待发货': col_map['pending_ship'] = i
        elif h == '待售后': col_map['pending_after_sale'] = i
        elif h == '在售中': col_map['on_sale'] = i
        elif h == '仓库中': col_map['in_warehouse'] = i
        elif h == '草稿箱': col_map['draft'] = i
        elif h == '24h需发货': col_map['ship_24h'] = i
        elif h == '超时未发货': col_map['overdue_ship'] = i
        elif h == '24h需售后': col_map['after_sale_24h'] = i
        elif h == '待整改': col_map['pending_rectify'] = i
        elif h == '最近违规时间': col_map['last_violation'] = i
        elif h == '工单': col_map['work_order'] = i
        elif h == '资金': col_map['fund'] = i
        elif h == '体验金': col_map['experience_fund'] = i
        elif h == '运费险价格': col_map['insurance_price'] = i
        elif h == '30天有效订单': col_map['valid_orders_30d'] = i
        elif h == '被清洗订单': col_map['cleaned_orders'] = i
        elif h == '违规数': col_map['violation_count'] = i
        elif h == '商品分': col_map['product_score'] = i
        elif h == '物流分': col_map['logistics_score'] = i
        elif h == '服务分': col_map['service_score'] = i
        elif h == '异常包裹': col_map['abnormal_parcel'] = i
        elif h == '回收站': col_map['recycle_bin'] = i
        elif h == '退款金额（今日支付）': col_map['refund_today'] = i
        elif h == '店铺ID': col_map['shop_id_col'] = i
        elif h == '千川ID': col_map['qianchuan_id'] = i
        elif h == '需开票': col_map['need_invoice'] = i
        elif h == '今订单': col_map['today_orders'] = i
        elif h == '分组': col_map['group_name'] = i
        elif h == '备注': col_map['remark'] = i

    def g(row, key, default=''):
        idx = col_map.get(key)
        if idx is not None and idx < len(row) and row[idx] is not None:
            return row[idx]
        return default

    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        shop_name = str(g(row, 'shop_name', '')).strip()
        if not shop_name:
            continue

        sid_col = g(row, 'shop_id_col', '')
        shop_id_str = str(sid_col).strip() if sid_col else ''
        group = str(g(row, 'group_name', '')).strip()

        existing_shop = conn.execute("SELECT id FROM shops WHERE shop_name=?", (shop_name,)).fetchone()
        if existing_shop:
            db_shop_id = existing_shop['id']
            conn.execute("UPDATE shops SET shop_id=?, group_name=? WHERE id=?", (shop_id_str, group, db_shop_id))
        else:
            cur = conn.execute("INSERT INTO shops(shop_name, shop_id, group_name) VALUES(?,?,?)", (shop_name, shop_id_str, group))
            db_shop_id = cur.lastrowid

        conn.execute("""INSERT OR REPLACE INTO daily_snapshot(
            shop_id, snapshot_date, status, expire_info, warning, amount, order_count,
            exposure, clicks, refund_amount, experience_score, pending_ship,
            pending_after_sale, on_sale, in_warehouse, draft, ship_24h, overdue_ship,
            after_sale_24h, pending_rectify, last_violation, work_order, fund,
            experience_fund, insurance_price, valid_orders_30d, cleaned_orders,
            violation_count, product_score, logistics_score, service_score,
            abnormal_parcel, recycle_bin, refund_today, qianchuan_id, need_invoice, today_orders, remark
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (db_shop_id, snapshot_date,
             str(g(row, 'status', '')),
             str(g(row, 'expire_info', '')),
             safe_int(g(row, 'warning', 0)),
             safe_float(g(row, 'amount', 0)),
             safe_int(g(row, 'order_count', 0)),
             safe_int(g(row, 'exposure', 0)),
             safe_int(g(row, 'clicks', 0)),
             safe_float(g(row, 'refund_amount', 0)),
             str(g(row, 'experience_score', '')),
             safe_int(g(row, 'pending_ship', 0)),
             safe_int(g(row, 'pending_after_sale', 0)),
             safe_int(g(row, 'on_sale', 0)),
             safe_int(g(row, 'in_warehouse', 0)),
             safe_int(g(row, 'draft', 0)),
             str(g(row, 'ship_24h', '')),
             safe_int(g(row, 'overdue_ship', 0)),
             str(g(row, 'after_sale_24h', '')),
             safe_int(g(row, 'pending_rectify', 0)),
             str(g(row, 'last_violation', '')),
             safe_int(g(row, 'work_order', 0)),
             safe_float(g(row, 'fund', 0)),
             safe_float(g(row, 'experience_fund', 0)),
             str(g(row, 'insurance_price', '')),
             safe_int(g(row, 'valid_orders_30d', 0)),
             safe_int(g(row, 'cleaned_orders', 0)),
             safe_int(g(row, 'violation_count', 0)),
             str(g(row, 'product_score', '')),
             str(g(row, 'logistics_score', '')),
             str(g(row, 'service_score', '')),
             safe_int(g(row, 'abnormal_parcel', 0)),
             safe_int(g(row, 'recycle_bin', 0)),
             safe_float(g(row, 'refund_today', 0)),
             str(g(row, 'qianchuan_id', '')),
             safe_float(g(row, 'need_invoice', 0)),
             safe_int(g(row, 'today_orders', 0)),
             str(g(row, 'remark', ''))))
        cnt += 1

    conn.commit()
    conn.close()
    wb.close()
    return cnt


def auto_import_all():
    for fpath, d in scan_excel_files():
        import_excel(fpath, d)


def get_summary(conn, snapshot_date):
    return dict(conn.execute("""
        SELECT COUNT(*) as shop_count,
               COALESCE(SUM(amount),0) as total_amount,
               COALESCE(SUM(order_count),0) as total_orders,
               COALESCE(SUM(refund_amount),0) as total_refund,
               COALESCE(SUM(refund_today),0) as total_refund_today,
               COALESCE(SUM(pending_ship),0) as total_pending_ship,
               COALESCE(SUM(overdue_ship),0) as total_overdue_ship,
               COALESCE(SUM(exposure),0) as total_exposure,
               COALESCE(SUM(clicks),0) as total_clicks,
               COALESCE(SUM(order_count),0) as total_today_orders
        FROM daily_snapshot WHERE snapshot_date=?
    """, (snapshot_date,)).fetchone())


def get_range_summary(conn, start_date, end_date):
    return dict(conn.execute("""
        SELECT COUNT(DISTINCT shop_id) as shop_count,
               COALESCE(SUM(amount),0) as total_amount,
               COALESCE(SUM(order_count),0) as total_orders,
               COALESCE(SUM(refund_amount),0) as total_refund,
               COALESCE(SUM(refund_today),0) as total_refund_today,
               COALESCE(SUM(pending_ship),0) as total_pending_ship,
               COALESCE(SUM(overdue_ship),0) as total_overdue_ship,
               COALESCE(SUM(exposure),0) as total_exposure,
               COALESCE(SUM(clicks),0) as total_clicks,
               COALESCE(SUM(order_count),0) as total_today_orders
        FROM daily_snapshot WHERE snapshot_date >= ? AND snapshot_date <= ?
    """, (start_date, end_date)).fetchone())


@app.get("/", response_class=HTMLResponse)
def index(request: Request, qdate: str = ""):
    conn = get_db()
    today = date.today().isoformat()
    auto_import_all()

    available_dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC LIMIT 60"
    ).fetchall()]

    if not available_dates:
        conn.close()
        return templates.TemplateResponse(request, "index.html", {
            "request": request, "summary": None, "shops": [], "today": today,
            "available_dates": [], "selected_date": today
        })

    selected_date = qdate if qdate in available_dates else available_dates[0]
    summary = get_summary(conn, selected_date)

    sd = date.fromisoformat(selected_date)
    summary_7d = get_range_summary(conn, (sd - timedelta(days=6)).isoformat(), selected_date)
    summary_15d = get_range_summary(conn, (sd - timedelta(days=14)).isoformat(), selected_date)
    summary_30d = get_range_summary(conn, (sd - timedelta(days=29)).isoformat(), selected_date)

    shops = [dict(r) for r in conn.execute("""
        SELECT s.shop_name, s.group_name, d.* FROM daily_snapshot d
        JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date=?
        ORDER BY d.amount DESC
    """, (selected_date,)).fetchall()]

    groups = sorted(set(s['group_name'] for s in shops if s['group_name']))

    conn.close()
    return templates.TemplateResponse(request, "index.html", {
        "request": request, "summary": summary, "shops": shops,
        "today": today, "available_dates": available_dates, "selected_date": selected_date,
        "groups": groups, "summary_7d": summary_7d, "summary_15d": summary_15d, "summary_30d": summary_30d
    })


@app.get("/api/summary/range_detail")
def api_summary_range_detail(start: str, end: str, metric: str):
    conn = get_db()
    metric_map = {
        'total_amount': 'SUM(d.amount)',
        'total_orders': 'SUM(d.order_count)',
        'total_today_orders': 'SUM(d.order_count)',
        'total_refund': 'SUM(d.refund_amount)',
        'total_refund_today': 'SUM(d.refund_today)',
        'total_pending_ship': 'SUM(d.pending_ship)',
        'total_overdue_ship': 'SUM(d.overdue_ship)',
        'total_exposure': 'SUM(d.exposure)',
        'total_clicks': 'SUM(d.clicks)',
    }
    col = metric_map.get(metric, 'SUM(d.amount)')
    shops = [dict(r) for r in conn.execute(f"""
        SELECT s.shop_name, {col} as val
        FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id
        WHERE d.snapshot_date >= ? AND d.snapshot_date <= ?
        GROUP BY d.shop_id HAVING val > 0 ORDER BY val DESC
    """, (start, end)).fetchall()]
    conn.close()
    return {"shops": shops, "metric": metric, "start": start, "end": end}


@app.get("/ranking", response_class=HTMLResponse)
def ranking_page(request: Request, sort_by: str = "amount", days: int = 1, group: str = ""):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    valid_sorts = {
        "amount": "SUM(d.amount) DESC",
        "orders": "SUM(d.order_count) DESC",
        "refund": "SUM(d.refund_amount) DESC",
        "exposure": "SUM(d.exposure) DESC",
        "clicks": "SUM(d.clicks) DESC",
    }
    order_clause = valid_sorts.get(sort_by, "SUM(d.amount) DESC")

    group_filter = ""
    params = [start_date, today]
    if group:
        group_filter = " AND s.group_name = ?"
        params.append(group)

    shops = [dict(r) for r in conn.execute(f"""
        SELECT s.shop_name, s.group_name,
               COUNT(*) as days_count,
               SUM(d.amount) as total_amount,
               SUM(d.order_count) as total_orders,
               SUM(d.refund_amount) as total_refund,
               SUM(d.exposure) as total_exposure,
               SUM(d.clicks) as total_clicks,
               SUM(d.today_orders) as total_today_orders,
               SUM(d.pending_ship) as total_pending_ship,
               SUM(d.overdue_ship) as total_overdue_ship
        FROM daily_snapshot d JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date >= ? AND d.snapshot_date <= ? {group_filter}
        GROUP BY d.shop_id
        ORDER BY {order_clause}
    """, params).fetchall()]

    all_groups = sorted(set(r['group_name'] for r in conn.execute(
        "SELECT DISTINCT s.group_name FROM shops s JOIN daily_snapshot d ON d.shop_id=s.id WHERE s.group_name != ''"
    ).fetchall()))

    conn.close()
    return templates.TemplateResponse(request, "ranking.html", {
        "request": request, "shops": shops, "sort_by": sort_by,
        "days": days, "today": today, "group": group, "all_groups": all_groups
    })


@app.get("/trends", response_class=HTMLResponse)
def trends_page(request: Request, days: int = 7):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    daily = [dict(r) for r in conn.execute("""
        SELECT snapshot_date,
               SUM(amount) as total_amount,
               SUM(order_count) as total_orders,
               SUM(refund_amount) as total_refund,
               SUM(exposure) as total_exposure,
               SUM(clicks) as total_clicks,
               SUM(pending_ship) as total_pending_ship,
               SUM(overdue_ship) as total_overdue_ship
        FROM daily_snapshot
        WHERE snapshot_date >= ? AND snapshot_date <= ?
        GROUP BY snapshot_date ORDER BY snapshot_date
    """, (start_date, today)).fetchall()]

    all_shops = [dict(r) for r in conn.execute("SELECT DISTINCT s.shop_name FROM shops s JOIN daily_snapshot d ON d.shop_id=s.id ORDER BY s.shop_name").fetchall()]

    conn.close()
    return templates.TemplateResponse(request, "trends.html", {
        "request": request, "daily": daily, "days": days, "today": today,
        "all_shops": all_shops
    })


@app.get("/alerts", response_class=HTMLResponse)
def alerts_page(request: Request):
    conn = get_db()
    auto_import_all()

    today = date.today().isoformat()
    available_dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC LIMIT 1"
    ).fetchall()]

    if not available_dates:
        conn.close()
        return templates.TemplateResponse(request, "alerts.html", {
            "request": request, "alerts": [], "today": today
        })

    selected_date = available_dates[0]

    shops = [dict(r) for r in conn.execute("""
        SELECT s.shop_name, s.group_name, d.* FROM daily_snapshot d
        JOIN shops s ON d.shop_id = s.id
        WHERE d.snapshot_date=?
    """, (selected_date,)).fetchall()]

    alerts = []
    for shop in shops:
        reasons = []
        if shop['overdue_ship'] > 0:
            reasons.append(f"超时未发货 {shop['overdue_ship']} 单")
        exp = safe_float(shop['experience_score'])
        if exp > 0 and exp < 70:
            reasons.append(f"体验分偏低: {shop['experience_score']}")
        if shop['violation_count'] > 0:
            reasons.append(f"违规 {shop['violation_count']} 次")
        if shop['pending_after_sale'] > 0:
            reasons.append(f"待售后 {shop['pending_after_sale']} 单")
        if shop['abnormal_parcel'] > 0:
            reasons.append(f"异常包裹 {shop['abnormal_parcel']} 个")
        if safe_float(shop['last_violation']):
            reasons.append(f"最近违规: {shop['last_violation']}")
        if reasons:
            alerts.append({"shop": shop, "reasons": reasons})

    alerts.sort(key=lambda x: len(x['reasons']), reverse=True)
    conn.close()
    return templates.TemplateResponse(request, "alerts.html", {
        "request": request, "alerts": alerts, "today": today
    })


@app.get("/api/summary/detail")
def api_summary_detail(date: str, metric: str):
    conn = get_db()
    shops = []
    if metric == 'total_orders':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.order_count as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.order_count>0 ORDER BY d.order_count DESC", (date,)).fetchall()]
    elif metric == 'total_today_orders':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.today_orders as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.today_orders>0 ORDER BY d.today_orders DESC", (date,)).fetchall()]
    elif metric == 'total_amount':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.amount as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.amount>0 ORDER BY d.amount DESC", (date,)).fetchall()]
    elif metric == 'total_refund':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.refund_amount as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.refund_amount>0 ORDER BY d.refund_amount DESC", (date,)).fetchall()]
    elif metric == 'total_refund_today':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.refund_today as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.refund_today>0 ORDER BY d.refund_today DESC", (date,)).fetchall()]
    elif metric == 'total_pending_ship':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.pending_ship as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.pending_ship>0 ORDER BY d.pending_ship DESC", (date,)).fetchall()]
    elif metric == 'total_overdue_ship':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.overdue_ship as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? AND d.overdue_ship>0 ORDER BY d.overdue_ship DESC", (date,)).fetchall()]
    elif metric == 'total_exposure':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.exposure as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? ORDER BY d.exposure DESC", (date,)).fetchall()]
    elif metric == 'total_clicks':
        shops = [dict(r) for r in conn.execute("SELECT s.shop_name, d.clicks as val FROM daily_snapshot d JOIN shops s ON d.shop_id=s.id WHERE d.snapshot_date=? ORDER BY d.clicks DESC", (date,)).fetchall()]
    conn.close()
    return {"shops": shops, "metric": metric, "date": date}


@app.get("/api/shop/detail")
def api_shop_detail(shop_name: str, range: str = "1"):
    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE shop_name=?", (shop_name,)).fetchone()
    if not shop:
        conn.close()
        return {"error": "not found"}
    days = int(range) if range.isdigit() else 1
    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()
    history = [dict(r) for r in conn.execute("""
        SELECT * FROM daily_snapshot WHERE shop_id=? AND snapshot_date>=? AND snapshot_date<=?
        ORDER BY snapshot_date
    """, (shop['id'], start_date, today)).fetchall()]
    conn.close()
    return {"shop": dict(shop), "history": history, "days": days}


@app.get("/api/dates")
def api_dates():
    conn = get_db()
    dates = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date"
    ).fetchall()]
    conn.close()
    return {"dates": dates}


@app.get("/shop/{shop_name}", response_class=HTMLResponse)
def shop_detail(request: Request, shop_name: str, days: int = 30):
    conn = get_db()
    auto_import_all()

    shop = conn.execute("SELECT * FROM shops WHERE shop_name=?", (shop_name,)).fetchone()
    if not shop:
        conn.close()
        return RedirectResponse("/", status_code=302)

    today = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()

    history = [dict(r) for r in conn.execute("""
        SELECT * FROM daily_snapshot
        WHERE shop_id=? AND snapshot_date >= ? AND snapshot_date <= ?
        ORDER BY snapshot_date
    """, (shop['id'], start_date, today)).fetchall()]

    latest = history[-1] if history else None
    conn.close()
    return templates.TemplateResponse(request, "shop_detail.html", {
        "request": request, "shop": dict(shop), "history": history,
        "latest": latest, "days": days, "today": today
    })


@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request, msg: str = ""):
    conn = get_db()
    rows = [r['snapshot_date'] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM daily_snapshot ORDER BY snapshot_date DESC"
    ).fetchall()]
    conn.close()
    grouped = {}
    for d in rows:
        y, m, _ = d.split('-')
        grouped.setdefault(y, {}).setdefault(m, []).append(d)
    return templates.TemplateResponse(request, "import.html", {
        "request": request, "grouped_dates": grouped, "msg": msg
    })


@app.post("/api/import")
async def api_import(file: UploadFile = File(...)):
    content = await file.read()
    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    import openpyxl
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    fname = file.filename or "upload.xlsx"
    wb.close()
    os.unlink(tmp_path)

    d = extract_date_from_filename(fname)
    if not d:
        d = date.today().isoformat()

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp2:
        tmp2.write(content)
        tmp2_path = tmp2.name

    cnt = import_excel(tmp2_path, d)
    os.unlink(tmp2_path)

    if cnt > 0:
        return RedirectResponse(f"/import?msg=成功导入 {cnt} 条店铺数据（{d}）", status_code=302)
    else:
        return RedirectResponse(f"/import?msg=该日期数据已存在或文件为空（{d}）", status_code=302)


if __name__ == "__main__":
    import uvicorn

    LAST_ACTIVITY = [time.time()]
    IDLE_TIMEOUT = 120

    @app.middleware("http")
    async def idle_timeout_middleware(request: Request, call_next):
        LAST_ACTIVITY[0] = time.time()
        response = await call_next(request)
        return response

    def check_idle():
        while True:
            time.sleep(10)
            if time.time() - LAST_ACTIVITY[0] > IDLE_TIMEOUT:
                print(f"空闲超过 {IDLE_TIMEOUT} 秒，自动关闭服务...")
                os._exit(0)

    t = threading.Thread(target=check_idle, daemon=True)
    t.start()

    uvicorn.run(app, host="0.0.0.0", port=8001)